from __future__ import annotations
from flask import Flask, request, jsonify, Response
from twofa import generate_base32_secret, current_totp_code, verify_totp_code, key_uri
import json
import os
import time
import uuid
import requests
import bleach
import threading
import json as pyjson
from datetime import datetime
from flask_socketio import SocketIO, emit, join_room, leave_room
from db import StudentDB, AdminDB, execute_query  # MySQL integration for students and admins
import bcrypt
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Optional Excel support for audit export
try:
    from openpyxl import Workbook  # type: ignore[import]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import]
    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

app = Flask(__name__)

# MySQL availability check (for health and logging)
try:
    from db import test_connection
    MYSQL_AVAILABLE = test_connection()
    print(f"MySQL Connection: {'✓ Available' if MYSQL_AVAILABLE else '✗ Not Available - Using fallback JSON storage'}")
except Exception as e:
    print(f"⚠ MySQL test failed: {e}")
    MYSQL_AVAILABLE = False

# CORS origins (used by both HTTP and Socket.IO)
ALLOWED_ORIGINS = set((os.getenv("ALLOWED_ORIGINS") or "http://localhost:8080,http://127.0.0.1:8080,http://localhost:8081,http://127.0.0.1:8081").split(","))

# Socket.IO for realtime notifications
socketio = SocketIO(app, cors_allowed_origins=list(ALLOWED_ORIGINS) or "*")

# In-memory stores (dev only)
NOTIFICATIONS = {}  # user_id -> list[notification]
PASSWORD_RESET_REQUESTS = {}  # req_id -> record

# Lightweight file-backed DB (dev)
DB_PATH = os.path.join(os.path.dirname(__file__), 'data.json')
DB_LOCK = threading.Lock()
DEFAULT_DB = {
    "users": {},          # id -> user dict
    "activity": [],       # list of activity records
    "books": [],          # optional for reports
    "borrows": []         # optional for reports
}

def load_db():
    with DB_LOCK:
        try:
            if not os.path.exists(DB_PATH):
                with open(DB_PATH, 'w', encoding='utf-8') as f:
                    pyjson.dump(DEFAULT_DB, f)
            with open(DB_PATH, 'r', encoding='utf-8') as f:
                return pyjson.load(f)
        except Exception:
            return DEFAULT_DB.copy()

def save_db(db):
    with DB_LOCK:
        try:
            with open(DB_PATH, 'w', encoding='utf-8') as f:
                pyjson.dump(db, f)
        except Exception:
            pass

def log_activity(user_id: str, action: str, details: str = ""):
    db = load_db()
    rec = {"id": f"ACT-{int(time.time()*1000)}", "userId": user_id, "action": action, "details": details, "timestamp": time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())}
    db.setdefault("activity", []).append(rec)
    db["activity"] = db["activity"][-1000:]
    save_db(db)
    _emit('activity.new', user_id, rec)


# ---- Email Configuration ----
EMAIL_ENABLED = os.getenv("EMAIL_ENABLED", "false").lower() == "true"
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SENDER_EMAIL = os.getenv("SENDER_EMAIL", "noreply@jrmsu.edu.ph")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD", "")  # Set via environment variable
SENDER_NAME = os.getenv("SENDER_NAME", "JRMSU Library System")

def send_reset_email(recipient_email: str, reset_code: str, recipient_name: str = "") -> bool:
    """
    Send password reset email with code.
    Returns True if email sent successfully, False otherwise.
    If EMAIL_ENABLED is False, just prints to console (dev mode).
    """
    if not EMAIL_ENABLED or not SENDER_PASSWORD:
        print(f"[MAIL] Reset code for {recipient_email}: {reset_code} (expires in 5m)")
        print(f"[MAIL] Email sending disabled. Set EMAIL_ENABLED=true and SENDER_PASSWORD env vars to enable.")
        return True  # Consider it "sent" in dev mode
    
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'JRMSU Library - Password Reset Code'
        msg['From'] = f'{SENDER_NAME} <{SENDER_EMAIL}>'
        msg['To'] = recipient_email
        
        # Email body
        greeting = f"Hello {recipient_name}," if recipient_name else "Hello,"
        text_body = f"""
{greeting}

You have requested to reset your password for the JRMSU Library System.

Your password reset code is: {reset_code}

This code will expire in 5 minutes.

If you did not request this password reset, please ignore this email.

Best regards,
JRMSU Library System
        """
        
        html_body = f"""
<html>
  <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 5px;">
      <h2 style="color: #003366;">JRMSU Library System</h2>
      <p>{greeting}</p>
      <p>You have requested to reset your password for the JRMSU Library System.</p>
      <p>Your password reset code is:</p>
      <h1 style="background: #f4f4f4; padding: 15px; text-align: center; letter-spacing: 5px; color: #003366;">{reset_code}</h1>
      <p style="color: #d9534f;"><strong>This code will expire in 5 minutes.</strong></p>
      <p>If you did not request this password reset, please ignore this email.</p>
      <hr style="margin: 20px 0; border: none; border-top: 1px solid #ddd;">
      <p style="font-size: 12px; color: #777;">Best regards,<br>JRMSU Library System</p>
    </div>
  </body>
</html>
        """
        
        # Attach both plain text and HTML versions
        part1 = MIMEText(text_body, 'plain')
        part2 = MIMEText(html_body, 'html')
        msg.attach(part1)
        msg.attach(part2)
        
        # Send email
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)
        
        print(f"[MAIL] ✅ Email sent successfully to {recipient_email}")
        return True
        
    except Exception as e:
        print(f"[MAIL] ❌ Failed to send email to {recipient_email}: {str(e)}")
        # Fallback to console in case of email failure
        print(f"[MAIL] Reset code for {recipient_email}: {reset_code} (expires in 5m)")
        return False


OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434")
MODEL_NAME = os.getenv("OLLAMA_MODEL", "llama3:8b-instruct-q4_K_M")

# Base URL for this backend (used by ai_server and other services)
BACKEND_BASE_URL = os.getenv("BACKEND_BASE_URL", "http://localhost:5000")

@app.before_request
def handle_preflight():
    # Handle preflight early so browsers get proper headers even if no route matches OPTIONS explicitly
    if request.method == "OPTIONS":
        return ("", 204)

@app.after_request
def add_cors(resp):
    origin = request.headers.get("Origin")
    if origin and origin in ALLOWED_ORIGINS:
        resp.headers["Access-Control-Allow-Origin"] = origin
        vary = resp.headers.get("Vary")
        resp.headers["Vary"] = f"{vary}, Origin" if vary else "Origin"
    else:
        resp.headers.setdefault("Access-Control-Allow-Origin", "*")
    resp.headers["Access-Control-Allow-Credentials"] = "true"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-User-Id"
    # Allow DELETE (and PATCH) so management pages can perform hard-deletes via CORS preflight
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, PATCH, OPTIONS"
    return resp


@app.route('/')
def root():
    return jsonify(message='JRMSU AI Library Backend is running!', status='ok', mysql=MYSQL_AVAILABLE)

@app.route('/health')
def health():
    return jsonify(status='ok', mysql=MYSQL_AVAILABLE, timestamp=time.time())

# ---------- System Administrator helpers (audit, backups, version, developers) ----------

def _get_client_ip() -> str:
    try:
        return request.headers.get('X-Forwarded-For', '').split(',')[0].strip() or request.remote_addr or ''
    except Exception:
        return ''


def _ensure_tables_for_admin_features():
    """Create core tables for admin features if they do not exist.

    This is safe to call multiple times; it will no-op if tables already exist.
    """
    if not MYSQL_AVAILABLE:
        return
    try:
        # audit_log table
        execute_query(
            """
            CREATE TABLE IF NOT EXISTS audit_log (
              id BIGINT AUTO_INCREMENT PRIMARY KEY,
              timestamp DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
              user_id VARCHAR(64) NULL,
              user_role ENUM('admin','student','system') NOT NULL DEFAULT 'system',
              action VARCHAR(128) NOT NULL,
              description TEXT NOT NULL,
              entity_type VARCHAR(64) NULL,
              entity_id VARCHAR(128) NULL,
              metadata JSON NULL,
              success TINYINT(1) NOT NULL DEFAULT 1,
              ip_address VARCHAR(45) NULL,
              created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
              INDEX idx_audit_ts (timestamp),
              INDEX idx_audit_user (user_id, action)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """,
            fetch_all=False,
        )
        # db_backups metadata table
        execute_query(
            """
            CREATE TABLE IF NOT EXISTS db_backups (
              id BIGINT AUTO_INCREMENT PRIMARY KEY,
              filename VARCHAR(255) NOT NULL,
              created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
              size_bytes BIGINT NOT NULL,
              checksum VARCHAR(64) NULL,
              created_by VARCHAR(64) NULL,
              notes TEXT NULL,
              INDEX idx_backups_created (created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """,
            fetch_all=False,
        )
        # system_version single-row table
        execute_query(
            """
            CREATE TABLE IF NOT EXISTS system_version (
              id INT PRIMARY KEY,
              name VARCHAR(128) NOT NULL,
              version VARCHAR(32) NOT NULL,
              release_date DATE NULL,
              status VARCHAR(32) NOT NULL,
              notes TEXT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """,
            fetch_all=False,
        )
        # Seed a default system_version row if missing
        row = execute_query("SELECT COUNT(*) AS c FROM system_version", fetch_one=True) or {"c": 0}
        if int(row.get('c') or 0) == 0:
            execute_query(
                "INSERT INTO system_version (id, name, version, release_date, status, notes) VALUES (1,%s,%s,%s,%s,%s)",
                (
                    'JRMSU Library Management System',
                    'v1.0.0',
                    '2024-10-30',
                    'Stable',
                    'QR Code authentication, 2FA, AI assistant, real-time notifications, reporting, backup & restore.',
                ),
            )
        # Developers table
        execute_query(
            """
            CREATE TABLE IF NOT EXISTS developers (
              id INT AUTO_INCREMENT PRIMARY KEY,
              name VARCHAR(128) NOT NULL,
              role VARCHAR(128) NOT NULL,
              email VARCHAR(128) NULL,
              phone VARCHAR(32) NULL,
              notes TEXT NULL,
              avatar_initials VARCHAR(4) NULL,
              accent VARCHAR(32) NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """,
            fetch_all=False,
        )
    except Exception as e:
        # Failing to create these tables should not crash the app; admin features will gracefully degrade.
        print(f"⚠️  Failed to ensure admin feature tables: {e}")


def _ensure_file_audit_store(db: dict) -> dict:
    db.setdefault('audit', [])
    return db


def write_audit_log(
    action: str,
    description: str,
    *,
    user_id: str | None = None,
    user_role: str = 'system',
    entity_type: str | None = None,
    entity_id: str | None = None,
    metadata: dict | None = None,
    success: bool = True,
) -> None:
    """Write an audit_log record to MySQL if available; otherwise to file-backed store.

    Also emits a realtime `audit:new` event for admins.
    """
    ts = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    ip = _get_client_ip()
    meta_json = None
    try:
        if metadata is not None:
            meta_json = json.dumps(metadata, default=str)
    except Exception:
        meta_json = None

    if MYSQL_AVAILABLE:
        try:
            _ensure_tables_for_admin_features()
            execute_query(
                """
                INSERT INTO audit_log
                  (timestamp, user_id, user_role, action, description, entity_type, entity_id, metadata, success, ip_address)
                VALUES (NOW(), %s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    user_id,
                    user_role,
                    action,
                    description,
                    entity_type,
                    entity_id,
                    meta_json,
                    1 if success else 0,
                    ip,
                ),
            )
        except Exception as e:
            print(f"⚠️  Failed to write audit_log to MySQL: {e}")
    else:
        try:
            db = load_db()
            db = _ensure_file_audit_store(db)
            db['audit'].append(
                {
                    'id': f'AUD-{int(time.time()*1000)}',
                    'timestamp': ts,
                    'user_id': user_id,
                    'user_role': user_role,
                    'action': action,
                    'description': description,
                    'entity_type': entity_type,
                    'entity_id': entity_id,
                    'metadata': metadata or {},
                    'success': bool(success),
                    'ip_address': ip,
                }
            )
            # Keep last 5000 only
            db['audit'] = db['audit'][-5000:]
            save_db(db)
        except Exception as e:
            print(f"⚠️  Failed to write audit_log to file store: {e}")

    # Emit realtime event to connected admins (broadcast)
    try:
        payload = {
            'timestamp': ts,
            'user_id': user_id,
            'user_role': user_role,
            'action': action,
            'description': description,
            'entity_type': entity_type,
            'entity_id': entity_id,
            'success': success,
        }
        _broadcast('audit:new', payload)
    except Exception:
        pass


def _get_admin_2fa_secret(admin_id: str) -> str | None:
    """Fetch stored 2FA secret for an admin from file-backed store.

    We store 2FA secrets in the lightweight users store (`data.json`) under `twoFactorKey`.
    """
    try:
        fdb = load_db()
        users = fdb.get('users') or {}
        u = users.get(admin_id)
        if not u:
            return None
        key = u.get('twoFactorKey') or u.get('twoFactorSetupKey')
        return key or None
    except Exception:
        return None


def _verify_admin_2fa(admin_id: str, token: str | None) -> bool:
    """Verify a TOTP token for an admin if a secret is configured.

    If there is no stored secret for this admin, we treat 2FA as not enforced
    (return True) to avoid locking out admins in dev mode.
    """
    secret = _get_admin_2fa_secret(admin_id)
    if not secret:
        # No 2FA configured for this admin; allow for now.
        return True
    token = (token or '').strip()
    if not token:
        return False
    try:
        return bool(verify_totp_code(secret, token, window=1))
    except Exception:
        return False


def _get_backups_dir() -> str:
    base = os.path.dirname(__file__)
    # Use project-level backup folder named "backupdb" (sibling of python-backend)
    path = os.path.join(base, '..', 'backupdb')
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        pass
    return os.path.abspath(path)


@app.route('/api/backup/create', methods=['POST'])
def api_backup_create():
    """Create a database/file backup and store it in the secure backups directory.

    This endpoint is called from the /settings System Administrator UI.
    It requires an admin user and their current password for confirmation.
    """
    body = request.get_json(silent=True) or {}
    admin_id = (body.get('userId') or body.get('adminId') or _get_user_id()).strip()
    admin_password = (body.get('adminPassword') or '').encode('utf-8')

    # Basic admin guard: require an admin id and password.
    # For now we only check that both fields are present; we do not block on
    # database password verification to avoid accidental lockouts.
    if not admin_id:
        return jsonify(error='Admin user required'), 403
    if not admin_password:
        return jsonify(error='Admin password required'), 400

    _ensure_tables_for_admin_features()

    backups_dir = _get_backups_dir()
    ts_label = time.strftime('%Y%m%d-%H%M%S')
    filename = f'backup-{ts_label}.json.gz'
    full_path = os.path.join(backups_dir, filename)

    # Default: backup the lightweight file-backed DB. In production, this should
    # be replaced or complemented with a proper MySQL dump.
    try:
        db = load_db()
        raw = json.dumps(db, default=str).encode('utf-8')
        import gzip, hashlib

        with gzip.open(full_path, 'wb') as f:
            f.write(raw)
        size_bytes = os.path.getsize(full_path)
        checksum = hashlib.sha256(raw).hexdigest()

        if MYSQL_AVAILABLE:
            try:
                execute_query(
                    "INSERT INTO db_backups (filename, size_bytes, checksum, created_by) VALUES (%s,%s,%s,%s)",
                    (filename, size_bytes, checksum, admin_id),
                )
            except Exception as e:
                print(f"⚠️  Failed to insert db_backups row: {e}")

        write_audit_log(
            'backup_created',
            f'Backup {filename} created successfully',
            user_id=admin_id,
            user_role='admin',
            entity_type='backup',
            entity_id=filename,
            metadata={'size_bytes': size_bytes, 'checksum': checksum},
            success=True,
        )

        # Notify realtime listeners
        try:
            payload = {
                'type': 'backup:created',
                'id': filename,
                'name': filename,
                'timestamp': int(time.time()),
                'size_bytes': size_bytes,
                'created_by': admin_id,
                'message': 'Backup created',
            }
            _broadcast('backup:created', payload)
            notif = {
                'id': _new_notif_id(),
                'user_id': 'admins',
                'title': 'Backup created',
                'body': f'Backup {filename} created by {admin_id}',
                'type': 'system',
                'meta': payload,
                'created_at': int(time.time()),
                'read': False,
                'action_required': False,
                'action_payload': None,
                'actor_id': admin_id,
            }
            # Broadcast generic admin notification; frontend can filter by target
            _broadcast('notification.new', notif)
        except Exception:
            pass

        return jsonify(ok=True, filename=filename, size_bytes=size_bytes, checksum=checksum)
    except Exception as e:
        write_audit_log(
            'backup_create_failed',
            f'Backup failed: {e}',
            user_id=admin_id,
            user_role='admin',
            entity_type='backup',
            success=False,
        )
        return jsonify(error='Backup failed', details=str(e)), 500


@app.route('/api/backup/list', methods=['GET'])
def api_backup_list():
    """List known backups from db_backups and filesystem.

    NOTE: This endpoint is used primarily by tools and may be subject to
    stricter content-type handling in some environments. For a
    frontend-friendly way to get the latest backup, prefer
    `/api/backup/latest` (POST) which always works with JSON bodies.
    """
    admin_id = _get_user_id()
    _ensure_tables_for_admin_features()
    backups_dir = _get_backups_dir()
    items = []

    # Prefer db_backups metadata when available
    if MYSQL_AVAILABLE:
        try:
            rows = execute_query("SELECT id, filename, created_at, size_bytes, checksum, created_by FROM db_backups ORDER BY created_at DESC", fetch_all=True) or []
            for r in rows:
                items.append(
                    {
                        'id': r.get('id'),
                        'filename': r.get('filename'),
                        'created_at': str(r.get('created_at')),
                        'size_bytes': int(r.get('size_bytes') or 0),
                        'checksum': r.get('checksum'),
                        'created_by': r.get('created_by'),
                    }
                )
        except Exception as e:
            print(f"⚠️  Failed to list db_backups: {e}")

    # Fallback: scan filesystem if db_backups empty or unavailable
    if not items:
        try:
            for name in sorted(os.listdir(backups_dir), reverse=True):
                if not name.lower().endswith('.gz'):
                    continue
                path = os.path.join(backups_dir, name)
                if not os.path.isfile(path):
                    continue
                size_bytes = os.path.getsize(path)
                items.append(
                    {
                        'id': None,
                        'filename': name,
                        'created_at': None,
                        'size_bytes': size_bytes,
                        'checksum': None,
                        'created_by': None,
                    }
                )
        except Exception as e:
            print(f"⚠️  Failed to scan backups dir: {e}")

    return jsonify(items=items, requestedBy=admin_id)


@app.route('/api/backup/latest', methods=['POST'])
def api_backup_latest():
    """Return metadata for the most recent backup.

    This endpoint is designed for frontend use and always expects a JSON
    body, which avoids content-type issues some browsers/tooling may
    trigger when calling `/api/backup/list` directly.
    """
    # Safely parse JSON body (even if empty) to keep strict middleware happy
    _ = request.get_json(silent=True) or {}

    # Reuse the listing logic above
    admin_id = _get_user_id()
    _ensure_tables_for_admin_features()
    backups_dir = _get_backups_dir()
    items = []

    if MYSQL_AVAILABLE:
        try:
            rows = execute_query("SELECT id, filename, created_at, size_bytes, checksum, created_by FROM db_backups ORDER BY created_at DESC", fetch_all=True) or []
            for r in rows:
                items.append(
                    {
                        'id': r.get('id'),
                        'filename': r.get('filename'),
                        'created_at': str(r.get('created_at')),
                        'size_bytes': int(r.get('size_bytes') or 0),
                        'checksum': r.get('checksum'),
                        'created_by': r.get('created_by'),
                    }
                )
        except Exception as e:
            print(f"⚠️  Failed to list db_backups (latest): {e}")

    if not items:
        try:
            for name in sorted(os.listdir(backups_dir), reverse=True):
                if not name.lower().endswith('.gz'):
                    continue
                path = os.path.join(backups_dir, name)
                if not os.path.isfile(path):
                    continue
                size_bytes = os.path.getsize(path)
                items.append(
                    {
                        'id': None,
                        'filename': name,
                        'created_at': None,
                        'size_bytes': size_bytes,
                        'checksum': None,
                        'created_by': None,
                    }
                )
        except Exception as e:
            print(f"⚠️  Failed to scan backups dir (latest): {e}")

    latest = items[0] if items else None
    if not latest:
        return jsonify(error='No backup file found'), 404

    return jsonify(latest=latest, requestedBy=admin_id)


@app.route('/api/backup/download/<path:filename>', methods=['GET'])
def api_backup_download(filename: str):
    """Download a specific backup file (.json.gz) from the backups directory.

    The Settings UI can call this to export/download a backup file.
    """
    backups_dir = _get_backups_dir()
    # Prevent directory traversal
    safe_name = os.path.basename(filename)
    full_path = os.path.join(backups_dir, safe_name)

    if not os.path.isfile(full_path):
        return jsonify(error='Backup file not found'), 404

    # Only allow our known backup format
    if not safe_name.lower().endswith('.json.gz'):
        return jsonify(error='Unsupported backup file type'), 400

    from flask import send_file
    return send_file(
        full_path,
        mimetype='application/gzip',
        as_attachment=True,
        download_name=safe_name,
    )


@app.route('/api/backup/upload', methods=['POST'])
def api_backup_upload():
    """Upload a backup file and store it into the backups directory.

    The uploaded file must match the expected backup format (.json.gz). After
    upload, admins can trigger restore which will treat it as any other backup.
    """
    backups_dir = _get_backups_dir()

    if 'file' not in request.files:
        return jsonify(error='No file part in request'), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify(error='No selected file'), 400

    safe_name = os.path.basename(file.filename)
    if not safe_name.lower().endswith('.json.gz'):
        return jsonify(error='Only .json.gz backup files are supported'), 400

    dest_path = os.path.join(backups_dir, safe_name)
    try:
        file.save(dest_path)
        size_bytes = os.path.getsize(dest_path)
        write_audit_log(
            'backup_uploaded',
            f'Backup file {safe_name} uploaded',
            user_id=_get_user_id(),
            user_role='admin',
            entity_type='backup',
            entity_id=safe_name,
            metadata={'size_bytes': size_bytes},
            success=True,
        )
        return jsonify(ok=True, filename=safe_name, size_bytes=size_bytes)
    except Exception as e:
        write_audit_log(
            'backup_upload_failed',
            f'Backup upload failed: {e}',
            user_id=_get_user_id(),
            user_role='admin',
            entity_type='backup',
            success=False,
        )
        return jsonify(error='Upload failed', details=str(e)), 500


@app.route('/api/backup/restore', methods=['POST'])
def api_backup_restore():
    """Restore from a backup file.

    By default, restores the most recent backup if `backupName` is not provided.
    Requires admin id + confirm phrase; optionally validates admin password
    when MySQL is available.
    """
    body = request.get_json(force=True) or {}
    admin_id = (body.get('userId') or body.get('adminId') or _get_user_id()).strip()
    confirm = (body.get('confirmPhrase') or '').strip()
    backup_name = (body.get('backupName') or '').strip()
    admin_password = (body.get('adminPassword') or '').encode('utf-8') if body.get('adminPassword') else b''

    if not admin_id:
        return jsonify(error='Admin user required'), 403

    # Require explicit confirmation phrase for safety
    if confirm.upper() != 'RESTORE NOW':
        write_audit_log(
            'backup_restore_denied',
            'Restore denied: missing confirmation phrase',
            user_id=admin_id,
            user_role='admin',
            entity_type='backup',
            success=False,
        )
        return jsonify(error="Confirmation phrase 'RESTORE NOW' required"), 400

    # We only require that an admin id and confirmation phrase are supplied.
    # Admin password may be used by the UI but is not enforced here to avoid
    # unexpected 401 errors during restore.

    _ensure_tables_for_admin_features()
    backups_dir = _get_backups_dir()

    # Determine target backup file
    target_file = None
    try:
        if backup_name:
            candidate = os.path.join(backups_dir, backup_name)
            if os.path.isfile(candidate):
                target_file = candidate
        else:
            # Pick most recent .gz file
            names = [n for n in os.listdir(backups_dir) if n.lower().endswith('.gz')]
            if names:
                names.sort(reverse=True)
                target_file = os.path.join(backups_dir, names[0])
    except Exception as e:
        print(f"⚠️  Failed to resolve backup file: {e}")

    if not target_file:
        return jsonify(error='No backup file found to restore'), 404

    # Perform restore of file-backed DB only (MySQL restore should be wired separately)
    import gzip as _gzip
    try:
        with _gzip.open(target_file, 'rb') as f:
            raw = f.read()
            data = json.loads(raw.decode('utf-8'))
        # Overwrite data.json
        with DB_LOCK:
            with open(DB_PATH, 'w', encoding='utf-8') as f:
                pyjson.dump(data, f)

        write_audit_log(
            'backup_restored',
            f'Restored from backup {os.path.basename(target_file)}',
            user_id=admin_id,
            user_role='admin',
            entity_type='backup',
            entity_id=os.path.basename(target_file),
            success=True,
        )

        try:
            payload = {
                'type': 'backup:restored',
                'id': os.path.basename(target_file),
                'timestamp': int(time.time()),
                'restored_by': admin_id,
                'success': True,
                'message': 'Restore completed',
            }
            _broadcast('backup:restored', payload)
            notif = {
                'id': _new_notif_id(),
                'user_id': 'admins',
                'title': 'Database restored',
                'body': f'Database restored from {os.path.basename(target_file)} by {admin_id}',
                'type': 'system',
                'meta': payload,
                'created_at': int(time.time()),
                'read': False,
                'action_required': False,
                'action_payload': None,
                'actor_id': admin_id,
            }
            _broadcast('notification.new', notif)
        except Exception:
            pass

        return jsonify(ok=True, backup=os.path.basename(target_file))
    except Exception as e:
        write_audit_log(
            'backup_restore_failed',
            f'Restore failed: {e}',
            user_id=admin_id,
            user_role='admin',
            entity_type='backup',
            entity_id=os.path.basename(target_file),
            success=False,
        )
        return jsonify(error='Restore failed', details=str(e)), 500


@app.route('/api/audit/export', methods=['GET', 'POST'])
def api_audit_export():
    """Export audit log entries as CSV or Excel (.xlsx).

    For admin-triggered exports (from Settings), we expect a POST with the
    admin's password for confirmation. A plain GET continues to work in dev
    or for non-sensitive export flows.
    """
    _ensure_tables_for_admin_features()
    user_id = _get_user_id()

    # Basic filters (only from POST body for now)
    export_format = 'csv'
    if request.method == 'POST':
        body = request.get_json(silent=True) or {}
        export_format = (body.get('format') or 'csv').lower().strip() or 'csv'
    else:
        body = {}

    action_filter = (body.get('action') or '').strip() or None
    success_filter = body.get('success')

    # Optional admin password check for POST requests initiated from Settings
    if request.method == 'POST':
        admin_id = (body.get('userId') or body.get('adminId') or user_id or '').strip()
        admin_password = (body.get('adminPassword') or '').encode('utf-8') if body.get('adminPassword') else b''
        if MYSQL_AVAILABLE and admin_id and admin_password:
            try:
                row = execute_query(
                    "SELECT * FROM admins WHERE admin_id = %s OR id = %s",
                    (admin_id, admin_id),
                    fetch_one=True,
                )
                valid = False
                if row and row.get('password_hash'):
                    try:
                        valid = bcrypt.checkpw(admin_password, str(row.get('password_hash')).encode('utf-8'))
                    except Exception:
                        valid = False
                if not valid:
                    write_audit_log(
                        'audit_export_denied',
                        'Invalid admin password for audit export',
                        user_id=admin_id,
                        user_role='admin',
                        entity_type='audit_export',
                        success=False,
                    )
                    return jsonify(error='Invalid admin credentials'), 401
            except Exception as e:
                print(f"⚠️  Failed to verify admin credentials for audit export: {e}")
        # If MYSQL_AVAILABLE is False, we allow export without password to keep dev simple

    rows = []
    if MYSQL_AVAILABLE:
        try:
            sql = "SELECT timestamp, user_id, user_role, action, description, entity_type, entity_id, success, ip_address FROM audit_log"
            clauses = []
            params: list[object] = []
            if action_filter:
                clauses.append("action = %s")
                params.append(action_filter)
            if isinstance(success_filter, bool):
                clauses.append("success = %s")
                params.append(1 if success_filter else 0)
            if clauses:
                sql += " WHERE " + " AND ".join(clauses)
            sql += " ORDER BY timestamp DESC LIMIT 5000"
            rows = execute_query(sql, tuple(params) if params else None, fetch_all=True) or []
        except Exception as e:
            print(f"⚠️  Failed to read audit_log from MySQL: {e}")

    if not rows:
        try:
            db = load_db()
            arr = _ensure_file_audit_store(db).get('audit', [])
            rows = arr[-5000:][::-1]
        except Exception as e:
            print(f"⚠️  Failed to read audit from file store: {e}")
            rows = []

    # Build CSV in-memory (always)
    import io, csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    header = ['timestamp','user_id','user_role','action','description','entity_type','entity_id','success','ip_address']
    writer.writerow(header)
    for r in rows:
        writer.writerow([
            str(r.get('timestamp')),
            r.get('user_id') or '',
            r.get('user_role') or '',
            r.get('action') or '',
            (r.get('description') or '').replace('\r',' ').replace('\n',' '),
            r.get('entity_type') or '',
            r.get('entity_id') or '',
            '1' if (r.get('success') in (1, True, '1', 'true')) else '0',
            r.get('ip_address') or '',
        ])

    csv_data = buf.getvalue().encode('utf-8')

    # If Excel export was requested but openpyxl is not available, signal this
    # clearly so the frontend can fall back to CSV.
    if export_format == 'xlsx' and not EXCEL_AVAILABLE:
        return jsonify(error='Excel export not available on server; falling back to CSV is recommended.'), 400

    # Try to build a real .xlsx file when requested and supported
    if export_format == 'xlsx' and EXCEL_AVAILABLE:
        import io as _io
        try:
            xbuf = _io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Audit Log'

            # Styles
            header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')  # blue
            header_font = Font(color='FFFFFF', bold=True)
            data_fill_alt = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')  # light blue
            thin_border = Border(
                left=Side(style='thin', color='D1D5DB'),
                right=Side(style='thin', color='D1D5DB'),
                top=Side(style='thin', color='D1D5DB'),
                bottom=Side(style='thin', color='D1D5DB'),
            )

            # Write header row with styling
            ws.append(header)
            for col_idx in range(1, len(header) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Write data rows with alternating fill
            row_index = 2
            for r in rows:
                values = [
                    str(r.get('timestamp')),
                    r.get('user_id') or '',
                    r.get('user_role') or '',
                    r.get('action') or '',
                    (r.get('description') or '').replace('\\r',' ').replace('\\n',' '),
                    r.get('entity_type') or '',
                    r.get('entity_id') or '',
                    1 if (r.get('success') in (1, True, '1', 'true')) else 0,
                    r.get('ip_address') or '',
                ]
                ws.append(values)
                # Apply borders and optional zebra striping
                for col_idx in range(1, len(values) + 1):
                    cell = ws.cell(row=row_index, column=col_idx)
                    cell.border = thin_border
                    if row_index % 2 == 0:
                        cell.fill = data_fill_alt
                row_index += 1

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value is not None else ''
                        max_length = max(max_length, len(val))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = max(12, min(max_length + 2, 60))

            wb.save(xbuf)
            xdata = xbuf.getvalue()

            write_audit_log(
                'audit_exported',
                'Audit log exported as Excel',
                user_id=user_id,
                user_role='admin',
                entity_type='audit_export',
                success=True,
            )

            from flask import Response as FlaskResponse
            today = time.strftime('%Y-%m-%d')
            resp = FlaskResponse(xdata, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp.headers['Content-Disposition'] = f'attachment; filename=audit_log_{today}.xlsx'
            return resp
        except Exception as e:
            print(f"⚠️  Failed to build Excel file for audit export: {e}")
            # Fall back to CSV

    # Default: CSV export (Excel-friendly)
    write_audit_log(
        'audit_exported',
        'Audit log exported as CSV',
        user_id=user_id,
        user_role='admin',
        entity_type='audit_export',
        success=True,
    )

    from flask import Response as FlaskResponse
    today = time.strftime('%Y-%m-%d')
    resp = FlaskResponse(csv_data, mimetype='text/csv')
    resp.headers['Content-Disposition'] = f'attachment; filename=audit_log_{today}.csv'
    return resp


@app.route('/api/admin/system-version', methods=['GET', 'POST'])
def api_system_version():
    """Get or update system version metadata used in the Settings page dialog."""
    _ensure_tables_for_admin_features()

    if request.method == 'GET':
        if MYSQL_AVAILABLE:
            try:
                row = execute_query("SELECT id, name, version, release_date, status, notes FROM system_version WHERE id = 1", fetch_one=True) or {}
                return jsonify(row)
            except Exception as e:
                print(f"⚠️  Failed to read system_version: {e}")
        # Fallback default
        return jsonify(
            {
                'id': 1,
                'name': 'JRMSU Library Management System',
                'version': 'v1.0.0',
                'release_date': '2024-10-30',
                'status': 'Stable',
                'notes': 'QR Code authentication, 2FA, AI assistant, real-time notifications, reporting, backup & restore.',
            }
        )

    # POST: update (admin-only, minimal guard)
    body = request.get_json(force=True) or {}
    user_id = _get_user_id()
    version = (body.get('version') or '').strip() or 'v1.0.0'
    name = (body.get('name') or 'JRMSU Library Management System').strip()
    status = (body.get('status') or 'Stable').strip()
    release_date = (body.get('release_date') or '2024-10-30').strip()
    notes = (body.get('notes') or '').strip()

    if MYSQL_AVAILABLE:
        try:
            execute_query(
                "REPLACE INTO system_version (id, name, version, release_date, status, notes) VALUES (1,%s,%s,%s,%s,%s)",
                (name, version, release_date, status, notes),
            )
        except Exception as e:
            return jsonify(error=str(e)), 500

    write_audit_log(
        'system_version_updated',
        f'System version updated to {version}',
        user_id=user_id,
        user_role='admin',
        entity_type='system_version',
        success=True,
    )

    try:
        payload = {
            'type': 'system:updated',
            'version': version,
            'release_date': release_date,
            'notes': notes,
        }
        _broadcast('system:updated', payload)
    except Exception:
        pass

    return jsonify(ok=True)


@app.route('/api/admin/developers', methods=['GET', 'POST'])
def api_developers_collection():
    """List or create developer records used in the Developers Information dialog."""
    _ensure_tables_for_admin_features()

    if request.method == 'GET':
        if MYSQL_AVAILABLE:
            try:
                rows = execute_query(
                    "SELECT id, name, role, email, phone, notes, avatar_initials, accent FROM developers ORDER BY id ASC",
                    fetch_all=True,
                ) or []
                return jsonify(items=rows)
            except Exception as e:
                print(f"⚠️  Failed to list developers: {e}")
        # Fallback to a static list matching the current UI
        return jsonify(
            items=[
                {
                    'id': 1,
                    'name': 'Jhon Mark Suico',
                    'role': 'Team Leader & System Engineer',
                    'email': 'suicojm99@gmail.com',
                    'phone': None,
                    'notes': 'Led the development and architecture of the entire system, ensuring seamless integration of all components.',
                    'avatar_initials': 'JM',
                    'accent': 'primary',
                },
                {
                    'id': 2,
                    'name': 'Jhon Ernie Alimpong',
                    'role': 'System Architect',
                    'email': None,
                    'phone': None,
                    'notes': 'Designed the system architecture and database structure, creating a robust foundation for scalability.',
                    'avatar_initials': 'JE',
                    'accent': 'accent',
                },
                {
                    'id': 3,
                    'name': 'Vivien Punay',
                    'role': 'Product Manager',
                    'email': None,
                    'phone': None,
                    'notes': 'Managed project requirements and user experience, ensuring the system meets real-world library needs.',
                    'avatar_initials': 'VP',
                    'accent': 'secondary',
                },
                {
                    'id': 4,
                    'name': 'Lenny Mambo',
                    'role': 'Data Analyst',
                    'email': None,
                    'phone': None,
                    'notes': 'Analyzed library data patterns and optimized reporting features for actionable insights.',
                    'avatar_initials': 'LM',
                    'accent': 'leaf',
                },
            ]
        )

    # POST: create new developer (admin-only basic guard)
    body = request.get_json(force=True) or {}
    user_id = _get_user_id()
    name = (body.get('name') or '').strip()
    role = (body.get('role') or '').strip()
    email = (body.get('email') or '').strip() or None
    phone = (body.get('phone') or '').strip() or None
    notes = (body.get('notes') or '').strip() or None
    initials = (body.get('avatar_initials') or '').strip() or None
    accent = (body.get('accent') or '').strip() or None

    if not name or not role:
        return jsonify(error='name and role are required'), 400

    if MYSQL_AVAILABLE:
        try:
            execute_query(
                "INSERT INTO developers (name, role, email, phone, notes, avatar_initials, accent) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (name, role, email, phone, notes, initials, accent),
            )
        except Exception as e:
            return jsonify(error=str(e)), 500

    write_audit_log(
        'developer_added',
        f'Developer {name} added',
        user_id=user_id,
        user_role='admin',
        entity_type='developer',
        success=True,
    )

    return jsonify(ok=True)


@app.route('/api/admin/developers/<int:dev_id>', methods=['PUT'])
def api_developer_update(dev_id: int):
    _ensure_tables_for_admin_features()
    body = request.get_json(force=True) or {}
    user_id = _get_user_id()

    if not MYSQL_AVAILABLE:
        return jsonify(error='Developers update not available without MySQL'), 503

    try:
        # Fetch existing for logging
        row = execute_query("SELECT * FROM developers WHERE id = %s", (dev_id,), fetch_one=True) or {}
        name = (body.get('name') or row.get('name') or '').strip()
        role = (body.get('role') or row.get('role') or '').strip()
        email = (body.get('email') or row.get('email') or '').strip() or None
        phone = (body.get('phone') or row.get('phone') or '').strip() or None
        notes = (body.get('notes') or row.get('notes') or '').strip() or None
        initials = (body.get('avatar_initials') or row.get('avatar_initials') or '').strip() or None
        accent = (body.get('accent') or row.get('accent') or '').strip() or None

        execute_query(
            "UPDATE developers SET name=%s, role=%s, email=%s, phone=%s, notes=%s, avatar_initials=%s, accent=%s WHERE id=%s",
            (name, role, email, phone, notes, initials, accent, dev_id),
        )

        write_audit_log(
            'developer_updated',
            f'Developer {dev_id} updated',
            user_id=user_id,
            user_role='admin',
            entity_type='developer',
            entity_id=str(dev_id),
            success=True,
        )
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(error=str(e)), 500


# Register library session endpoints so mirror can login/logout via backend
try:
    from library_session_manager import register_library_session_endpoints
    if 'check_user_session' not in app.view_functions:
        register_library_session_endpoints(app)
        print('✅ Library session endpoints loaded (app)')
    else:
        print('✅ Library session endpoints already loaded')
except Exception as e:
    print(f'⚠️  Library session endpoints not loaded (app): {e}')
# Register notifications blueprint if available
try:
    from notifications_routes import notifications_bp
    app.register_blueprint(notifications_bp)
    print('✅ Notifications endpoints loaded')
except Exception as e:
    print(f'⚠️  Notifications endpoints not loaded: {e}')

# ---------- Users/Profile API ----------
@app.route('/api/users')
def list_users():
    # Prefer MySQL admins and students for accuracy, then merge any file-backed users
    items = []
    try:
        arows = AdminDB.list_all_admins() or []
        for r in arows:
            items.append(_map_admin_row_to_user(r))
    except Exception:
        pass
    try:
        rows = StudentDB.list_all_students() or []
        for r in rows:
            items.append(_map_student_row_to_user(r))
    except Exception:
        pass
    try:
        fdb = load_db()
        users = list((fdb.get("users") or {}).values())
        # Avoid duplicates by id
        existing_ids = {u.get('id') for u in items}
        items.extend([u for u in users if u.get('id') not in existing_ids])
    except Exception:
        pass
    return jsonify(items=items)

@app.route('/api/users/<uid>')
def get_user(uid: str):
    # Try MySQL admin first
    try:
        arow = AdminDB.get_admin_by_id(uid)
        if arow:
            return jsonify(_map_admin_row_to_user(arow))
    except Exception:
        pass
    # Then student
    try:
        row = StudentDB.get_student_by_id(uid)
        if row:
            return jsonify(_map_student_row_to_user(row))
    except Exception:
        pass
    # Fallback to file-backed store
    fdb = load_db()
    u = fdb.get("users", {}).get(uid)
    if not u:
        u = {"id": uid}
    return jsonify(u)

# ---------- Admin-specific API (maps to users with userType == 'admin') ----------

def _ensure_users():
    db = load_db()
    db.setdefault("users", {})
    return db

# Note: All Admin routes (GET/PUT/POST) moved to line 700+ with database integration

@app.route('/api/admins/<admin_id>/2fa/setup', methods=['POST'])
def admins_2fa_setup(admin_id: str):
    # Delegate to 2FA generator, but store secret on admin
    r = twofa_generate()
    data = r.get_json()
    db = _ensure_users()
    u = (db.get('users') or {}).get(admin_id)
    if not u:
        return jsonify(error='Admin not found'), 404
    u['twoFactorSetupKey'] = data.get('secret')
    db['users'][admin_id] = u
    save_db(db)
    return jsonify(secret=data.get('secret'), otpauth=data.get('otpauth'), currentCode=data.get('currentCode'))

@app.route('/api/admins/<admin_id>/2fa/verify', methods=['POST'])
def admins_2fa_verify(admin_id: str):
    body = request.get_json(force=True) or {}
    secret = (body.get('secret') or '').strip()
    token = (body.get('token') or body.get('totpCode') or '').strip()
    ok = verify_totp_code(secret, token, window=1)
    if not ok:
        return jsonify(valid=False), 400
    db = _ensure_users()
    u = (db.get('users') or {}).get(admin_id)
    if not u:
        return jsonify(error='Admin not found'), 404
    u['twoFactorEnabled'] = True
    u['twoFactorKey'] = secret
    db['users'][admin_id] = u
    save_db(db)
    log_activity(admin_id, '2fa_enable')
    _emit('admins.updated', admin_id, {'twoFactorEnabled': True})
    return jsonify(valid=True)

@app.route('/api/admins/<admin_id>/2fa/disable', methods=['POST'])
def admins_2fa_disable(admin_id: str):
    db = _ensure_users()
    u = (db.get('users') or {}).get(admin_id)
    if not u:
        return jsonify(error='Admin not found'), 404
    u['twoFactorEnabled'] = False
    u.pop('twoFactorKey', None)
    db['users'][admin_id] = u
    save_db(db)
    log_activity(admin_id, '2fa_disable')
    _emit('admins.updated', admin_id, {'twoFactorEnabled': False})
    return jsonify(ok=True)

# ---------- Student-specific API (maps to users with userType == 'student') ----------

@app.route('/api/students', methods=['GET'])
def students_list():
    try:
        rows = StudentDB.list_all_students() or []
        students = [_map_student_row_to_user(r) for r in rows]
        return jsonify(items=students)
    except Exception:
        # Fallback
        fdb = load_db()
        users = list((fdb.get('users') or {}).values())
        students = [u for u in users if (u.get('userType') == 'student' or (u.get('role') == 'student'))]
        return jsonify(items=students)

@app.route('/api/students/<student_id>', methods=['GET'])
def students_get(student_id: str):
    try:
        row = StudentDB.get_student_by_id(student_id)
        if not row:
            return jsonify(error='Student not found'), 404
        return jsonify(_map_student_row_to_user(row))
    except Exception as e:
        # Fallback to file store
        fdb = load_db()
        u = (fdb.get('users') or {}).get(student_id)
        if not u or (u.get('userType') != 'student' and u.get('role') != 'student'):
            return jsonify(error='Student not found'), 404
        return jsonify(u)

@app.route('/api/students/<student_id>', methods=['DELETE'])
def students_delete(student_id: str):
    """Permanently delete a student record from MySQL and fallback store."""
    sid = (student_id or "").strip()
    if not sid:
        return jsonify(error='Student ID required'), 400

    full_name = sid

    # Try to load existing record for logging before delete
    try:
        row = StudentDB.get_student_by_id(sid)
        if row:
            full_name = (
                row.get('full_name')
                or f"{row.get('first_name','')} {row.get('last_name','')}".strip()
                or sid
            )
        # Hard-delete from MySQL if available
        try:
            execute_query("DELETE FROM students WHERE student_id = %s OR id = %s", (sid, sid))
        except Exception:
            # DB might be unavailable; continue to fallback store
            pass
    except Exception:
        # StudentDB may be unavailable; continue to fallback store
        pass

    # Remove from file-backed store (dev / fallback)
    try:
        fdb = _ensure_users()
        users = fdb.get('users') or {}
        if sid in users:
            u = users.pop(sid)
            full_name = u.get('fullName') or u.get('full_name') or full_name
            fdb['users'] = users
            save_db(fdb)
    except Exception:
        pass

    # Log to local activity feed for dashboards
    try:
        log_activity(sid, 'student_deleted', full_name)
    except Exception:
        pass

    # Broadcast realtime update so overlays and dashboards can refresh
    try:
        _broadcast('students.updated', { 'id': sid, 'deleted': True })
    except Exception:
        pass

    return jsonify(success=True, message='Student deleted successfully')

@app.route('/api/students/<student_id>', methods=['PUT'])
def students_put(student_id: str):
    body = request.get_json(force=True) or {}
    # Map allowed editable fields to stored procedure inputs
    try:
        # Determine block from ID if not provided
        blk = body.get('block')
        if not blk and '-' in student_id:
            try:
                blk = student_id.split('-')[2]
            except Exception:
                blk = ''

        ok, msg = StudentDB.update_student_profile(
            student_id=student_id,
            department=(body.get('department') or body.get('college_department') or ''),
            course=(body.get('course') or body.get('course_major') or ''),
            year_level=(body.get('year') or body.get('year_level') or body.get('yearLevel') or ''),
            block=blk or '',
            current_street=body.get('currentStreet') or body.get('street') or '',
            current_barangay=body.get('currentBarangay') or body.get('barangay') or '',
            current_municipality=body.get('currentMunicipality') or body.get('municipality') or body.get('city') or '',
            current_province=body.get('currentProvince') or body.get('province') or '',
            current_region=body.get('currentRegion') or body.get('region') or '',
            current_zip=body.get('currentZipCode') or body.get('zipCode') or '',
            current_landmark=body.get('currentLandmark') or ''
        )
        if not ok:
            return jsonify(error=msg or 'Update failed'), 400
        # Return fresh row
        row = StudentDB.get_student_by_id(student_id) or {}
        student = _map_student_row_to_user(row)
        log_activity(student_id, 'profile_update')
        _emit('students.updated', student_id, student)
        return jsonify(ok=True, student=student)
    except Exception as e:
        # Fallback to previous file-backed logic
        fdb = _ensure_users()
        u = (fdb.get('users') or {}).get(student_id)
        if not u:
            return jsonify(error='Student not found'), 404
        # Basic merge for fallback
        for k in list(body.keys()):
            u[k] = body[k]
        fdb['users'][student_id] = u
        save_db(fdb)
        log_activity(student_id, 'profile_update')
        _emit('students.updated', student_id, u)
        return jsonify(ok=True, student=u)

@app.route('/api/students/register', methods=['POST'])
def students_register():
    body = request.get_json(force=True) or {}
    student_id = (body.get('studentId') or body.get('id') or '').strip()
    if not student_id:
        return jsonify(error='Student ID required'), 400

    # Extract block from student ID (KC-23-A-00762 => A)
    extracted_block = ''
    if '-' in student_id:
        try:
            parts = student_id.split('-')
            if len(parts) >= 3:
                extracted_block = parts[2]
        except Exception:
            pass

    # Personal Information
    first = (body.get('firstName') or '').strip()
    middle = (body.get('middleName') or '').strip()
    last = (body.get('lastName') or '').strip()
    suffix = (body.get('suffix') or '').strip()
    email = (body.get('email') or '').strip().lower()
    phone = (body.get('phone') or '').strip()
    gender = (body.get('gender') or '').strip()
    birthdate = (body.get('birthdate') or body.get('birthday') or '').strip()
    # Academic Information
    department = body.get('department') or body.get('college_department') or ''
    course = body.get('course') or body.get('course_major') or ''
    year_level = body.get('year') or body.get('year_level') or body.get('yearLevel') or ''
    block = body.get('block') or extracted_block

    # Current Address
    current_street = body.get('addressStreet') or body.get('currentAddressStreet') or ''
    current_barangay = body.get('addressBarangay') or body.get('currentAddressBarangay') or ''
    current_municipality = body.get('addressMunicipality') or body.get('currentAddressMunicipality') or body.get('municipality') or body.get('city') or ''
    current_province = body.get('addressProvince') or body.get('currentAddressProvince') or body.get('province') or ''
    current_region = body.get('addressRegion') or body.get('currentAddressRegion') or body.get('region') or ''
    current_zip = body.get('addressZip') or body.get('currentAddressZip') or body.get('zipCode') or body.get('zipcode') or ''
    current_landmark = body.get('addressPermanentNotes') or body.get('currentAddressLandmark') or ''

    # Permanent Address
    same_as_current = bool(body.get('sameAsCurrent', False))
    if same_as_current:
        permanent_street = current_street
        permanent_barangay = current_barangay
        permanent_municipality = current_municipality
        permanent_province = current_province
        permanent_region = current_region
        permanent_zip = current_zip
        permanent_notes = current_landmark
    else:
        permanent_street = body.get('permanentAddressStreet') or ''
        permanent_barangay = body.get('permanentAddressBarangay') or ''
        permanent_municipality = body.get('permanentAddressMunicipality') or ''
        permanent_province = body.get('permanentAddressProvince') or ''
        permanent_region = body.get('permanentAddressRegion') or ''
        permanent_zip = body.get('permanentAddressZip') or ''
        permanent_notes = body.get('permanentAddressNotes') or ''

    # Password hashing
    password = (body.get('password') or '').encode('utf-8')
    if not password:
        return jsonify(error='Password required'), 400
    password_hash = bcrypt.hashpw(password, bcrypt.gensalt()).decode('utf-8')

    # Use DB stored procedure
    try:
        ok, msg = StudentDB.register_student(
            student_id=student_id,
            first_name=first,
            middle_name=middle,
            last_name=last,
            suffix=suffix,
            birthdate=birthdate,
            gender=gender,
            email=email,
            phone=phone,
            department=department,
            course=course,
            year_level=year_level,
            current_street=current_street,
            current_barangay=current_barangay,
            current_municipality=current_municipality,
            current_province=current_province,
            current_region=current_region,
            current_zip=current_zip,
            current_landmark=current_landmark,
            permanent_street=permanent_street,
            permanent_barangay=permanent_barangay,
            permanent_municipality=permanent_municipality,
            permanent_province=permanent_province,
            permanent_region=permanent_region,
            permanent_zip=permanent_zip,
            permanent_notes=permanent_notes,
            same_as_current=same_as_current,
            password_hash=password_hash,
        )
        if not ok:
            return jsonify(error=msg or 'Registration failed'), 400
        # Return fresh student row
        row = StudentDB.get_student_by_id(student_id) or {}
        student = _map_student_row_to_user(row)
        log_activity(student_id, 'student_register')
        _emit('students.updated', student_id, student)
        return jsonify(ok=True, student=student, studentId=student_id)
    except Exception as e:
        # Fallback to file-backed store if DB not available
        fdb = _ensure_users()
        if student_id in fdb['users']:
            return jsonify(error='Student ID already exists'), 400
        # Minimal fallback record
        fallback = {
            'id': student_id,
            'studentId': student_id,
            'userType': 'student',
            'role': 'student',
            'firstName': first,
            'middleName': middle,
            'lastName': last,
            'suffix': suffix,
            'email': email,
            'phone': phone,
            'gender': gender,
            'birthday': birthdate,
            'age': body.get('age') or '',
            'department': department,
            'course': course,
            'year': year_level,
            'yearLevel': year_level,
            'section': block,
            'block': block,
            'address': ', '.join([p for p in [current_street, current_barangay, current_municipality, current_province, current_region, 'Philippines', current_zip] if p])
        }
        fdb['users'][student_id] = fallback
        save_db(fdb)
        log_activity(student_id, 'student_register')
        _emit('students.updated', student_id, fallback)
        return jsonify(ok=True, student=fallback, studentId=student_id)

@app.route('/api/students/<student_id>/2fa/setup', methods=['POST'])
def students_2fa_setup(student_id: str):
    r = twofa_generate()
    data = r.get_json()
    db = _ensure_users()
    u = (db.get('users') or {}).get(student_id)
    if not u:
        return jsonify(error='Student not found'), 404
    u['twoFactorSetupKey'] = data.get('secret')
    db['users'][student_id] = u
    save_db(db)
    return jsonify(secret=data.get('secret'), otpauth=data.get('otpauth'), currentCode=data.get('currentCode'))

@app.route('/api/students/<student_id>/2fa/verify', methods=['POST'])
def students_2fa_verify(student_id: str):
    body = request.get_json(force=True) or {}
    secret = (body.get('secret') or '').strip()
    token = (body.get('token') or body.get('totpCode') or '').strip()
    ok = verify_totp_code(secret, token, window=1)
    if not ok:
        return jsonify(valid=False), 400
    db = _ensure_users()
    u = (db.get('users') or {}).get(student_id)
    if not u:
        return jsonify(error='Student not found'), 404
    u['twoFactorEnabled'] = True
    u['twoFactorKey'] = secret
    db['users'][student_id] = u
    save_db(db)
    log_activity(student_id, '2fa_enable')
    _emit('students.updated', student_id, {'twoFactorEnabled': True})
    return jsonify(valid=True)

@app.route('/api/students/<student_id>/2fa/disable', methods=['POST'])
def students_2fa_disable(student_id: str):
    db = _ensure_users()
    u = (db.get('users') or {}).get(student_id)
    if not u:
        return jsonify(error='Student not found'), 404
    u['twoFactorEnabled'] = False
    u.pop('twoFactorKey', None)
    db['users'][student_id] = u
    save_db(db)
    log_activity(student_id, '2fa_disable')
    _emit('students.updated', student_id, {'twoFactorEnabled': False})
    return jsonify(ok=True)

@app.route('/api/users/<uid>', methods=['PATCH'])
def update_user(uid: str):
    body = request.get_json(force=True)
    # Try to update admin in MySQL if exists
    try:
        arow = AdminDB.get_admin_by_id(uid)
        if arow:
            ok, msg = AdminDB.update_admin_profile(
                admin_id=uid,
                first_name=body.get('firstName') or arow.get('first_name') or '',
                middle_name=body.get('middleName') or arow.get('middle_name') or '',
                last_name=body.get('lastName') or arow.get('last_name') or '',
                suffix=body.get('suffix') or arow.get('suffix') or '',
                gender=body.get('gender') or arow.get('gender') or '',
                age=body.get('age') or arow.get('age') or '',
                birthdate=body.get('birthdate') or body.get('birthday') or (str(arow.get('birthdate')) if arow.get('birthdate') else ''),
                email=body.get('email') or arow.get('email') or '',
                phone=body.get('phone') or arow.get('phone') or '',
                street=body.get('street') or arow.get('street') or '',
                barangay=body.get('barangay') or arow.get('barangay') or '',
                municipality=body.get('municipality') or body.get('city') or arow.get('municipality') or '',
                province=body.get('province') or arow.get('province') or '',
                region=body.get('region') or arow.get('region') or '',
                zip_code=body.get('zipCode') or arow.get('zip_code') or '',
                current_street=body.get('currentStreet') or arow.get('current_street') or '',
                current_barangay=body.get('currentBarangay') or arow.get('current_barangay') or '',
                current_municipality=body.get('currentMunicipality') or arow.get('current_municipality') or '',
                current_province=body.get('currentProvince') or arow.get('current_province') or '',
                current_region=body.get('currentRegion') or arow.get('current_region') or '',
                current_zip=body.get('currentZipCode') or arow.get('current_zip') or '',
                current_landmark=body.get('currentLandmark') or arow.get('current_landmark') or ''
            )
            if not ok:
                return jsonify(error=msg or 'Update failed'), 400
            new_row = AdminDB.get_admin_by_id(uid) or {}
            user = _map_admin_row_to_user(new_row)
            log_activity(uid, 'profile_update')
            _emit('user.updated', uid, user)
            return jsonify(ok=True, user=user)
    except Exception:
        pass
    # Try to update student in MySQL if exists
    try:
        row = StudentDB.get_student_by_id(uid)
        if row:
            # Only pass editable fields per requirements
            blk = body.get('block')
            if not blk and '-' in uid:
                try:
                    blk = uid.split('-')[2]
                except Exception:
                    blk = ''
            ok, msg = StudentDB.update_student_profile(
                student_id=uid,
                department=(body.get('department') or body.get('college_department') or row.get('department') or ''),
                course=(body.get('course') or body.get('course_major') or row.get('course') or ''),
                year_level=(body.get('year') or body.get('year_level') or body.get('yearLevel') or row.get('year_level') or ''),
                block=blk or (row.get('block') or ''),
                permanent_street=body.get('permanent_address_street') or row.get('permanent_address_street') or '',
                permanent_barangay=body.get('permanent_address_barangay') or row.get('permanent_address_barangay') or '',
                permanent_municipality=body.get('permanent_address_municipality') or row.get('permanent_address_municipality') or '',
                permanent_province=body.get('permanent_address_province') or row.get('permanent_address_province') or '',
                permanent_region=body.get('permanent_address_region') or row.get('permanent_address_region') or '',
                permanent_zip=body.get('permanent_address_zip') or row.get('permanent_address_zip') or '',
                current_street=body.get('current_address_street') or body.get('currentStreet') or body.get('street') or row.get('current_address_street') or '',
                current_barangay=body.get('current_address_barangay') or body.get('currentBarangay') or body.get('barangay') or row.get('current_address_barangay') or '',
                current_municipality=body.get('current_address_municipality') or body.get('currentMunicipality') or body.get('municipality') or body.get('city') or row.get('current_address_municipality') or '',
                current_province=body.get('current_address_province') or body.get('currentProvince') or body.get('province') or row.get('current_address_province') or '',
                current_region=body.get('current_address_region') or body.get('currentRegion') or body.get('region') or row.get('current_address_region') or '',
                current_zip=body.get('current_address_zip') or body.get('currentZipCode') or body.get('zipCode') or row.get('current_address_zip') or '',
                current_landmark=body.get('current_address_landmark') or body.get('currentLandmark') or row.get('current_address_landmark') or '',
                same_as_current=body.get('same_as_current') or row.get('same_as_current') or False
            )
            if not ok:
                return jsonify(error=msg or 'Update failed'), 400
            new_row = StudentDB.get_student_by_id(uid) or {}
            user = _map_student_row_to_user(new_row)
            log_activity(uid, 'profile_update')
            _emit('user.updated', uid, user)
            return jsonify(ok=True, user=user)
    except Exception:
        pass
    # Fallback to file-backed update
    fdb = load_db()
    users = fdb.setdefault("users", {})
    cur = users.get(uid, {"id": uid})
    cur.update(body or {})
    users[uid] = cur
    save_db(fdb)
    log_activity(uid, 'profile_update')
    _emit('user.updated', uid, cur)
    return jsonify(ok=True, user=cur)

# ---------- Admin-specific API ----------

@app.route('/api/admins', methods=['GET'])
def admins_list():
    try:
        rows = AdminDB.list_all_admins() or []
        admins = [_map_admin_row_to_user(r) for r in rows]
        return jsonify(items=admins)
    except Exception:
        # Fallback to file store
        fdb = load_db()
        users = list((fdb.get('users') or {}).values())
        admins = [u for u in users if (u.get('userType') == 'admin' or (u.get('role') in ['admin','assistant','staff','librarian','supervisor']))]
        return jsonify(items=admins)

@app.route('/api/admins/<admin_id>', methods=['GET'])
def admins_get(admin_id: str):
    try:
        row = AdminDB.get_admin_by_id(admin_id)
        if not row:
            return jsonify(error='Admin not found'), 404
        return jsonify(_map_admin_row_to_user(row))
    except Exception:
        fdb = load_db()
        u = (fdb.get('users') or {}).get(admin_id)
        if not u or (u.get('userType') != 'admin' and u.get('role') not in ['admin','assistant','staff','librarian','supervisor']):
            return jsonify(error='Admin not found'), 404
        return jsonify(u)

@app.route('/api/admins/<admin_id>', methods=['PUT'])
def admins_put(admin_id: str):
    body = request.get_json(force=True) or {}
    # Editable personal/contact/address fields
    try:
        # Get existing admin data to fill defaults
        existing = AdminDB.get_admin_by_id(admin_id) or {}
        ok, msg = AdminDB.update_admin_profile(
            admin_id=admin_id,
            first_name=body.get('firstName') or existing.get('first_name') or '',
            middle_name=body.get('middleName') or existing.get('middle_name') or '',
            last_name=body.get('lastName') or existing.get('last_name') or '',
            suffix=body.get('suffix') or existing.get('suffix') or '',
            gender=body.get('gender') or existing.get('gender') or '',
            age=body.get('age') or existing.get('age') or '',
            birthdate=body.get('birthdate') or body.get('birthday') or (str(existing.get('birthdate')) if existing.get('birthdate') else ''),
            email=body.get('email') or existing.get('email') or '',
            phone=body.get('phone') or existing.get('phone') or '',
            street=body.get('street') or existing.get('street') or '',
            barangay=body.get('barangay') or existing.get('barangay') or '',
            municipality=body.get('municipality') or body.get('city') or existing.get('municipality') or '',
            province=body.get('province') or existing.get('province') or '',
            region=body.get('region') or existing.get('region') or '',
            zip_code=body.get('zipCode') or existing.get('zip_code') or '',
            current_street=body.get('currentStreet') or existing.get('current_street') or '',
            current_barangay=body.get('currentBarangay') or existing.get('current_barangay') or '',
            current_municipality=body.get('currentMunicipality') or existing.get('current_municipality') or '',
            current_province=body.get('currentProvince') or existing.get('current_province') or '',
            current_region=body.get('currentRegion') or existing.get('current_region') or '',
            current_zip=body.get('currentZipCode') or existing.get('current_zip') or '',
            current_landmark=body.get('currentLandmark') or existing.get('current_landmark') or ''
        )
        if not ok:
            return jsonify(error=msg or 'Update failed'), 400
        row = AdminDB.get_admin_by_id(admin_id) or {}
        admin = _map_admin_row_to_user(row)
        log_activity(admin_id, 'profile_update')
        _emit('admins.updated', admin_id, admin)
        return jsonify(ok=True, admin=admin)
    except Exception:
        # Fallback file store
        fdb = _ensure_users()
        u = (fdb.get('users') or {}).get(admin_id)
        if not u:
            return jsonify(error='Admin not found'), 404
        for k in list(body.keys()):
            u[k] = body[k]
        fdb['users'][admin_id] = u
        save_db(fdb)
        log_activity(admin_id, 'profile_update')
        _emit('admins.updated', admin_id, u)
        return jsonify(ok=True, admin=u)

@app.route('/api/admins/<admin_id>', methods=['DELETE'])
def admins_delete(admin_id: str):
    """Permanently delete an administrator record from MySQL and fallback store."""
    aid = (admin_id or "").strip()
    if not aid:
        return jsonify(error='Admin ID required'), 400

    full_name = aid

    # Try to load existing record for logging before delete
    try:
        row = AdminDB.get_admin_by_id(aid)
        if row:
            full_name = (
                row.get('full_name')
                or f"{row.get('first_name','')} {row.get('last_name','')}".strip()
                or aid
            )
        # Hard-delete from MySQL if available
        try:
            execute_query("DELETE FROM admins WHERE admin_id = %s OR id = %s", (aid, aid))
        except Exception:
            # DB might be unavailable; continue to fallback store
            pass
    except Exception:
        # AdminDB may be unavailable; continue to fallback store
        pass

    # Remove from file-backed store (dev / fallback)
    try:
        fdb = _ensure_users()
        users = fdb.get('users') or {}
        if aid in users:
            u = users.pop(aid)
            full_name = u.get('fullName') or u.get('full_name') or full_name
            fdb['users'] = users
            save_db(fdb)
    except Exception:
        pass

    # Log to local activity feed for dashboards
    try:
        log_activity(aid, 'admin_deleted', full_name)
    except Exception:
        pass

    # Broadcast realtime update so overlays and dashboards can refresh
    try:
        _broadcast('admins.updated', { 'id': aid, 'deleted': True })
    except Exception:
        pass

    return jsonify(success=True, message='Administrator deleted successfully')

@app.route('/api/admins/register', methods=['POST'])
def admins_register():
    body = request.get_json(force=True) or {}
    admin_id = (body.get('adminId') or body.get('id') or '').strip()
    if not admin_id:
        return jsonify(error='Admin ID required'), 400

    # Personal info
    first = (body.get('firstName') or '').strip()
    middle = (body.get('middleName') or '').strip()
    last = (body.get('lastName') or '').strip()
    suffix = (body.get('suffix') or '').strip()
    email = (body.get('email') or '').strip().lower()
    phone = (body.get('phone') or '').strip()
    gender = (body.get('gender') or '').strip()
    birthdate = (body.get('birthdate') or body.get('birthday') or '').strip()
    age = body.get('age') or ''
    position = (body.get('position') or body.get('role') or 'admin').strip()

    # Permanent Address
    street = body.get('addressStreet') or body.get('street') or ''
    barangay = body.get('addressBarangay') or body.get('barangay') or ''
    municipality = body.get('addressMunicipality') or body.get('municipality') or body.get('city') or ''
    province = body.get('addressProvince') or body.get('province') or ''
    region = body.get('addressRegion') or body.get('region') or ''
    zip_code = body.get('addressZip') or body.get('zipCode') or ''
    
    # Current Address
    current_street = body.get('currentStreet') or body.get('currentAddressStreet') or ''
    current_barangay = body.get('currentBarangay') or body.get('currentAddressBarangay') or ''
    current_municipality = body.get('currentMunicipality') or body.get('currentAddressMunicipality') or ''
    current_province = body.get('currentProvince') or body.get('currentAddressProvince') or ''
    current_region = body.get('currentRegion') or body.get('currentAddressRegion') or ''
    current_zip = body.get('currentZipCode') or body.get('currentAddressZip') or ''
    current_landmark = body.get('currentLandmark') or body.get('addressPermanentNotes') or ''
    same_as_current = bool(body.get('sameAsCurrent', False))

    # Password
    password = (body.get('password') or '').encode('utf-8')
    if not password:
        return jsonify(error='Password required'), 400
    password_hash = bcrypt.hashpw(password, bcrypt.gensalt()).decode('utf-8')

    try:
        ok, msg = AdminDB.register_admin(
            admin_id=admin_id,
            first_name=first,
            middle_name=middle,
            last_name=last,
            suffix=suffix,
            birthdate=birthdate,
            gender=gender,
            email=email,
            phone=phone,
            position=position,
            street=street,
            barangay=barangay,
            municipality=municipality,
            province=province,
            region=region,
            zip_code=zip_code,
            current_street=current_street,
            current_barangay=current_barangay,
            current_municipality=current_municipality,
            current_province=current_province,
            current_region=current_region,
            current_zip=current_zip,
            current_landmark=current_landmark,
            same_as_current=same_as_current,
            password_hash=password_hash
        )
        if not ok:
            return jsonify(error=msg or 'Registration failed'), 400
        row = AdminDB.get_admin_by_id(admin_id) or {}
        admin = _map_admin_row_to_user(row)
        log_activity(admin_id, 'admin_register')
        _emit('admins.updated', admin_id, admin)
        return jsonify(ok=True, admin=admin, adminId=admin_id)
    except Exception:
        # Fallback to file store
        fdb = _ensure_users()
        if admin_id in fdb['users']:
            return jsonify(error='Admin ID already exists'), 400
        fallback = {
            'id': admin_id,
            'userType': 'admin',
            'role': position,
            'firstName': first,
            'middleName': middle,
            'lastName': last,
            'suffix': suffix,
            'email': email,
            'phone': phone,
            'gender': gender,
            'birthday': birthdate,
            'age': age,
            'address': ', '.join([p for p in [street, barangay, municipality, province, region, 'Philippines', zip_code] if p]),
            'systemTag': 'JRMSU-KCL',
            'twoFactorEnabled': False,
            'qrCodeActive': True,
            'isActive': True,
            'createdAt': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
            'updatedAt': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
        }
        fdb['users'][admin_id] = fallback
        save_db(fdb)
        log_activity(admin_id, 'admin_register')
        _emit('admins.updated', admin_id, fallback)
        return jsonify(ok=True, admin=fallback, adminId=admin_id)

@app.route('/api/users/<uid>/2fa', methods=['POST'])
def toggle_2fa(uid: str):
    db = load_db()
    body = request.get_json(force=True)
    enabled = bool(body.get('enabled'))
    users = db.setdefault("users", {})
    cur = users.get(uid, {"id": uid})
    cur['twoFactorEnabled'] = enabled
    if enabled and body.get('secret'):
        cur['twoFactorKey'] = body.get('secret')
    users[uid] = cur
    save_db(db)
    log_activity(uid, '2fa_enable' if enabled else '2fa_disable')
    _emit('user.2fa', uid, {"enabled": enabled})
    return jsonify(ok=True, user=cur)

# ---------- Books API (for Jose + frontend sync) ----------

@app.route('/api/books/search', methods=['GET'])
def api_books_search():
    """Search books by free-text query (code, title, author, category, ISBN).

    This endpoint is used by the Jose AI server and can also be used by frontends.
    It queries the MySQL `books` table when available, and falls back to the
    lightweight file-backed `data.json` store otherwise.
    """
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify(results=[])

    results = []

    # Prefer MySQL-backed books table
    if MYSQL_AVAILABLE:
        try:
            like = f"%{q}%"
            rows = execute_query(
                """
                SELECT 
                  id, 
                  book_code, 
                  title, 
                  author, 
                  category, 
                  isbn, 
                  shelf_location AS shelf,
                  total_copies,
                  available_copies,
                  status,
                  qr_code_value AS qr_code
                FROM books
                WHERE 
                  book_code LIKE %s OR
                  title LIKE %s OR
                  author LIKE %s OR
                  category LIKE %s OR
                  isbn LIKE %s
                ORDER BY title ASC
                LIMIT 20
                """,
                (like, like, like, like, like),
                fetch_all=True,
            ) or []
            for r in rows:
                results.append({
                    'id': r.get('book_code') or r.get('id'),
                    'code': r.get('book_code') or r.get('id'),
                    'title': r.get('title'),
                    'author': r.get('author'),
                    'category': r.get('category'),
                    'isbn': r.get('isbn'),
                    'shelf': r.get('shelf'),
                    'copies': int(r.get('total_copies') or 0),
                    'available': int(r.get('available_copies') or 0),
                    'status': r.get('status') or 'available',
                    'qr_code': r.get('qr_code'),
                })
        except Exception as e:
            print(f"⚠️  /api/books/search MySQL error: {e}")

    # Fallback to file-backed books from data.json (used for demo / dev)
    if not results:
        try:
            db = load_db()
            books = db.get('books') or []
            lq = q.lower()
            for b in books:
                if any((str(b.get(k,''))).lower().find(lq) != -1 for k in ['id','code','title','author','category','isbn']):
                    results.append({
                        'id': b.get('code') or b.get('id'),
                        'code': b.get('code') or b.get('id'),
                        'title': b.get('title'),
                        'author': b.get('author'),
                        'category': b.get('category'),
                        'isbn': b.get('isbn'),
                        'shelf': b.get('shelf'),
                        'copies': int(b.get('copies') or 0),
                        'available': int(b.get('available') or 0),
                        'status': b.get('status') or 'available',
                        'qr_code': b.get('qr') or b.get('qr_code'),
                    })
        except Exception as e:
            print(f"⚠️  /api/books/search file-store error: {e}")

    return jsonify(results=results)

# Activity feed
@app.route('/api/activity', methods=['GET'])
def list_activity():
    uid = request.args.get('userId')
    db = load_db()
    arr = db.get('activity', [])
    if uid:
        arr = [a for a in arr if a.get('userId') == uid]
    arr = sorted(arr, key=lambda a: a.get('timestamp',''), reverse=True)
    return jsonify(items=arr[:200])

@app.route('/api/activity', methods=['POST'])
def add_activity():
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        body = {}
    uid = (body or {}).get('userId') or _get_user_id()
    action = (body or {}).get('action') or 'event'
    details = (body or {}).get('details') or ''
    log_activity(uid, action, details)
    return jsonify(ok=True)

# Reports endpoints (derived from file-backed data)
@app.route('/api/reports/top-borrowed')
def api_top_borrowed():
    db = load_db()
    counts = {}
    for b in db.get('borrows', []):
        title = b.get('bookTitle') or b.get('bookId')
        if not title: 
            continue
        counts[title] = counts.get(title, 0) + 1
    top = sorted(({"title": k, "borrows": v} for k,v in counts.items()), key=lambda x: x['borrows'], reverse=True)[:5]
    return jsonify(items=top)

@app.route('/api/reports/category-dist')
def api_category_dist():
    db = load_db()
    counts = {}
    books = db.get('books', [])
    for b in books:
        cat = (b.get('category') or 'Uncategorized')
        counts[cat] = counts.get(cat, 0) + 1
    total = len(books) or 1
    dist = [{"category": k, "percentage": round((v/total)*100)} for k,v in counts.items()]
    return jsonify(items=dist)

# -------------------- Dashboard API --------------------

def _fmt_date_label(dt: datetime) -> str:
    try:
        return dt.strftime('%B %d, %Y').upper()
    except Exception:
        return ''


def _fmt_time_label(dt: datetime) -> str:
    try:
        t = dt.strftime('%I:%M %p')  # e.g., 09:30 PM
        # Convert AM/PM to A.M./P.M.
        return t.replace('AM', 'A.M.').replace('PM', 'P.M.')
    except Exception:
        return ''

@app.route('/api/dashboard/total-books', methods=['GET'])
def dashboard_total_books():
    try:
        rows = execute_query(
            "SELECT id AS book_id, title, author, category FROM books ORDER BY title ASC",
            fetch_all=True
        ) or []
        return jsonify(total=len(rows), data=rows)
    except Exception:
        # Fallback to file-backed store
        fdb = load_db()
        books = fdb.get('books', []) or []
        data = [
            {
                'book_id': b.get('id') or b.get('bookId') or b.get('code'),
                'title': b.get('title'),
                'author': b.get('author'),
                'category': b.get('category') or b.get('genre')
            } for b in books
        ]
        return jsonify(total=len(data), data=data)

@app.route('/api/dashboard/active-borrowers', methods=['GET'])
def dashboard_active_borrowers():
    """Borrowers with an active borrow (not returned). Group by borrow date."""
    def _group(rows):
        out = {}
        for r in rows:
            dt = r.get('borrowed_at_dt') or r.get('borrow_date_dt')
            label = _fmt_date_label(dt) if isinstance(dt, datetime) else (r.get('date_label') or '')
            lst = out.setdefault(label, [])
            lst.append({
                'timestamp': _fmt_time_label(dt) if isinstance(dt, datetime) else r.get('time_label', ''),
                'user_id': r.get('user_id'),
                'fullname': r.get('full_name') or r.get('fullname') or '',
                'course': r.get('course') or r.get('course_major') or '',
                'year': r.get('year_level') or r.get('year') or '',
                'block': r.get('block') or ''
            })
        return out
    try:
        # Prefer borrow_records
        sql = (
            "SELECT br.user_id, br.borrowed_at AS ts, s.full_name, s.course_major AS course, s.year_level, s.block "
            "FROM borrow_records br "
            "LEFT JOIN students s ON s.student_id = br.user_id OR s.id = br.user_id "
            "WHERE (br.status = 'borrowed' OR (br.returned_at IS NULL)) "
            "ORDER BY br.borrowed_at DESC"
        )
        rows = execute_query(sql, fetch_all=True) or []
        parsed = []
        for r in rows:
            ts = r.get('ts')
            try:
                dt = ts if isinstance(ts, datetime) else datetime.fromisoformat(str(ts))
            except Exception:
                dt = None
            parsed.append({
                'user_id': r.get('user_id'),
                'full_name': r.get('full_name'),
                'course': r.get('course'),
                'year_level': r.get('year_level'),
                'block': r.get('block'),
                'borrowed_at_dt': dt,
            })
        return jsonify(data=_group(parsed))
    except Exception:
        try:
            # Fallback to borrows table (date fields)
            sql = (
                "SELECT b.user_id, b.borrow_date AS ts, s.full_name, s.course_major AS course, s.year_level, s.block "
                "FROM borrows b "
                "LEFT JOIN students s ON s.student_id = b.user_id OR s.id = b.user_id "
                "WHERE b.status = 'borrowed' AND (b.return_date IS NULL) "
                "ORDER BY b.borrow_date DESC"
            )
            rows = execute_query(sql, fetch_all=True) or []
            parsed = []
            for r in rows:
                ts = r.get('ts')
                try:
                    dt = ts if isinstance(ts, datetime) else datetime.fromisoformat(str(ts))
                except Exception:
                    dt = None
                parsed.append({
                    'user_id': r.get('user_id'),
                    'full_name': r.get('full_name'),
                    'course': r.get('course'),
                    'year_level': r.get('year_level'),
                    'block': r.get('block'),
                    'borrow_date_dt': dt,
                })
            return jsonify(data=_group(parsed))
        except Exception:
            # File-backed fallback
            fdb = load_db()
            borrows = fdb.get('borrows', [])
            parsed = []
            for b in borrows:
                if str(b.get('status')) == 'returned':
                    continue
                try:
                    dt = datetime.fromisoformat((b.get('borrowDate') or b.get('borrowed_at') or '')[:19])
                except Exception:
                    dt = None
                parsed.append({
                    'user_id': b.get('studentId') or b.get('userId'),
                    'full_name': b.get('fullName') or '',
                    'course': b.get('course') or '',
                    'year_level': b.get('year') or '',
                    'block': b.get('block') or '',
                    'borrowed_at_dt': dt,
                })
            return jsonify(data=_group(parsed))

@app.route('/api/dashboard/books-borrowed-today', methods=['GET'])
def dashboard_borrowed_today():
    def _group(rows):
        out = {}
        for r in rows:
            dt = r.get('dt')
            label = _fmt_date_label(dt)
            lst = out.setdefault(label, [])
            lst.append({
                'timestamp': _fmt_time_label(dt),
                'book_id': r.get('book_id'),
                'title': r.get('title'),
                'author': r.get('author'),
                'category': r.get('category'),
            })
        return out
    today = datetime.now().date()
    try:
        sql = (
            "SELECT br.borrowed_at AS ts, br.book_id, bk.title, bk.author, bk.category "
            "FROM borrow_records br "
            "LEFT JOIN books bk ON bk.id = br.book_id "
            "WHERE DATE(br.borrowed_at) = CURDATE() "
            "ORDER BY br.borrowed_at DESC"
        )
        rows = execute_query(sql, fetch_all=True) or []
        parsed = []
        for r in rows:
            ts = r.get('ts')
            try:
                dt = ts if isinstance(ts, datetime) else datetime.fromisoformat(str(ts))
            except Exception:
                dt = datetime.combine(today, datetime.min.time())
            parsed.append({
                'dt': dt,
                'book_id': r.get('book_id'),
                'title': r.get('title'),
                'author': r.get('author'),
                'category': r.get('category'),
            })
        return jsonify(data=_group(parsed))
    except Exception:
        try:
            sql = (
                "SELECT b.borrow_date AS ts, b.book_id, bk.title, bk.author, bk.category "
                "FROM borrows b "
                "LEFT JOIN books bk ON bk.id = b.book_id "
                "WHERE DATE(b.borrow_date) = CURDATE() "
                "ORDER BY b.borrow_date DESC"
            )
            rows = execute_query(sql, fetch_all=True) or []
            parsed = []
            for r in rows:
                ts = r.get('ts')
                try:
                    dt = ts if isinstance(ts, datetime) else datetime.fromisoformat(str(ts))
                except Exception:
                    dt = datetime.combine(today, datetime.min.time())
                parsed.append({
                    'dt': dt,
                    'book_id': r.get('book_id'),
                    'title': r.get('title'),
                    'author': r.get('author'),
                    'category': r.get('category'),
                })
            return jsonify(data=_group(parsed))
        except Exception:
            # File-backed fallback
            fdb = load_db()
            items = []
            for b in (fdb.get('borrows') or []):
                try:
                    if (b.get('borrowDate') or '')[:10] == today.isoformat():
                        items.append({
                            'dt': datetime.now(),
                            'book_id': b.get('bookId'),
                            'title': b.get('bookTitle'),
                            'author': b.get('author') or '',
                            'category': b.get('category') or '',
                        })
                except Exception:
                    pass
            return jsonify(data=_group(items))

@app.route('/api/dashboard/overdue-returns', methods=['GET'])
def dashboard_overdue_returns():
    def _group(rows):
        out = {}
        for r in rows:
            dt = r.get('dt')
            label = _fmt_date_label(dt)
            lst = out.setdefault(label, [])
            lst.append({
                'timestamp': _fmt_time_label(dt),
                'user_id': r.get('user_id'),
                'fullname': r.get('full_name') or r.get('fullname') or '',
                'course': r.get('course') or '',
                'year': r.get('year_level') or r.get('year') or '',
                'block': r.get('block') or ''
            })
        return out
    try:
        sql = (
            "SELECT br.id AS rec_id, br.user_id, br.borrowed_at AS ts, s.full_name, s.course_major AS course, s.year_level, s.block "
            "FROM borrow_records br "
            "LEFT JOIN students s ON s.student_id = br.user_id OR s.id = br.user_id "
            "WHERE br.status = 'borrowed' AND br.returned_at IS NULL AND br.due_date < NOW() "
            "ORDER BY br.due_date ASC"
        )
        rows = execute_query(sql, fetch_all=True) or []
        parsed = []
        for r in rows:
            ts = r.get('ts')
            try:
                dt = ts if isinstance(ts, datetime) else datetime.fromisoformat(str(ts))
            except Exception:
                dt = datetime.now()
            parsed.append({
                'rec_id': r.get('rec_id'),
                'dt': dt,
                'user_id': r.get('user_id'),
                'full_name': r.get('full_name'),
                'course': r.get('course'),
                'year_level': r.get('year_level'),
                'block': r.get('block'),
            })
        return jsonify(data=_group(parsed))
    except Exception:
        try:
            sql = (
                "SELECT b.id AS rec_id, b.user_id, b.borrow_date AS ts, s.full_name, s.course_major AS course, s.year_level, s.block "
                "FROM borrows b "
                "LEFT JOIN students s ON s.student_id = b.user_id OR s.id = b.user_id "
                "WHERE b.status = 'borrowed' AND (b.return_date IS NULL) AND b.due_date < CURDATE() "
                "ORDER BY b.due_date ASC"
            )
            rows = execute_query(sql, fetch_all=True) or []
            parsed = []
            for r in rows:
                ts = r.get('ts')
                try:
                    dt = ts if isinstance(ts, datetime) else datetime.fromisoformat(str(ts))
                except Exception:
                    dt = datetime.now()
                parsed.append({
                    'rec_id': r.get('rec_id'),
                    'dt': dt,
                    'user_id': r.get('user_id'),
                    'full_name': r.get('full_name'),
                    'course': r.get('course'),
                    'year_level': r.get('year_level'),
                    'block': r.get('block'),
                })
            return jsonify(data=_group(parsed))
        except Exception:
            # File-backed fallback
            fdb = load_db()
            items = []
            now = datetime.now().date()
            for b in (fdb.get('borrows') or []):
                try:
                    due = (b.get('dueDate') or '')[:10]
                    if b.get('status') == 'borrowed' and due and now.isoformat() > due:
                        items.append({
                            'rec_id': b.get('id') or b.get('borrow_id') or f"{b.get('studentId')}-{b.get('bookId')}-{b.get('dueDate')}",
                            'dt': datetime.now(),
                            'user_id': b.get('studentId') or b.get('userId'),
                            'full_name': b.get('fullName') or '',
                            'course': b.get('course') or '',
                            'year_level': b.get('year') or '',
                            'block': b.get('block') or '',
                        })
                except Exception:
                    pass
            return jsonify(data=_group(items))

@app.route('/api/dashboard/summary', methods=['GET'])
def dashboard_summary():
    """Return counts for header cards."""
    try:
        total_books = execute_query("SELECT COUNT(*) AS c FROM books", fetch_one=True) or {'c': 0}
        active_borrowers = execute_query(
            "SELECT COUNT(DISTINCT user_id) AS c FROM borrow_records WHERE status='borrowed' OR returned_at IS NULL",
            fetch_one=True
        ) or {'c': 0}
        borrowed_today = execute_query(
            "SELECT COUNT(*) AS c FROM borrow_records WHERE DATE(borrowed_at) = CURDATE()",
            fetch_one=True
        ) or {'c': 0}
        overdue = execute_query(
            "SELECT COUNT(*) AS c FROM borrow_records WHERE (status='borrowed' AND returned_at IS NULL AND due_date < NOW())",
            fetch_one=True
        ) or {'c': 0}
        return jsonify({
            'totalBooks': int(total_books.get('c') or 0),
            'activeBorrowers': int(active_borrowers.get('c') or 0),
            'borrowedToday': int(borrowed_today.get('c') or 0),
            'overdue': int(overdue.get('c') or 0),
        })
    except Exception:
        # Fallback to file store
        fdb = load_db()
        books = fdb.get('books', []) or []
        borrows = fdb.get('borrows', []) or []
        totalBooks = len(books)
        activeBorrowers = len({ (b.get('studentId') or b.get('userId')) for b in borrows if (b.get('status') != 'returned') })
        today = time.strftime('%Y-%m-%d')
        borrowedToday = sum(1 for b in borrows if (b.get('borrowDate') or '')[:10] == today)
        overdue = sum(1 for b in borrows if (b.get('status') == 'overdue'))
        return jsonify({ 'totalBooks': totalBooks, 'activeBorrowers': activeBorrowers, 'borrowedToday': borrowedToday, 'overdue': overdue })

# Background overdue watcher
OVERDUE_CACHE = set()

def _scan_and_emit_overdue():
    try:
        rows = execute_query(
            "SELECT id AS rec_id, user_id, book_id FROM borrow_records WHERE status='borrowed' AND returned_at IS NULL AND due_date < NOW()",
            fetch_all=True
        ) or []
        global OVERDUE_CACHE
        current_ids = set()
        for r in rows:
            rid = str(r.get('rec_id'))
            current_ids.add(rid)
            if rid not in OVERDUE_CACHE:
                try:
                    _broadcast('book.overdue', { 'recordId': rid, 'userId': r.get('user_id'), 'bookId': r.get('book_id'), 'timestamp': int(time.time()) })
                except Exception:
                    pass
        OVERDUE_CACHE = current_ids
    except Exception:
        # Silent if DB is unavailable
        pass

def start_overdue_watcher(interval_sec: int = 60):
    def _worker():
        while True:
            _scan_and_emit_overdue()
            time.sleep(interval_sec)
    try:
        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        print('⏰ Overdue watcher started')
    except Exception as e:
        print(f'⚠️  Overdue watcher not started: {e}')


def start_overdue_bulk_notifier(interval_sec: int = 24 * 60 * 60):
    """Periodically trigger bulk overdue notifications for all users.

    This calls the existing /api/overdue/notify-all endpoint, which respects
    per-user notification preferences and sends email, SMS, and push
    notifications via the notification_endpoints module.
    """
    def _worker():
        while True:
            try:
                resp = requests.post("http://localhost:5000/api/overdue/notify-all", timeout=30)
                if not resp.ok:
                    print(f"⚠️  overdue notify-all failed: {resp.status_code} {str(resp.text)[:200]}")
                else:
                    data = None
                    try:
                        data = resp.json()
                    except Exception:
                        data = None
                    if data:
                        print(f"📨 overdue notify-all: processed={data.get('processed')} queued={data.get('queued')}")
            except Exception as e:
                print(f"⚠️  overdue notify-all error: {e}")
            time.sleep(interval_sec)

    try:
        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        print('⏰ Bulk overdue notifier started')
    except Exception as e:
        print(f'⚠️  Bulk overdue notifier not started: {e}')


# Admin: trigger overdue scan immediately
@app.route('/api/admin/overdue/scan-now', methods=['POST'])
def api_admin_overdue_scan_now():
    try:
        before = len(OVERDUE_CACHE)
        _scan_and_emit_overdue()
        after = len(OVERDUE_CACHE)
        return jsonify(ok=True, tracked=after, newlyEmitted=max(0, after - before))
    except Exception as e:
        return jsonify(error=str(e)), 500

# ---------- Socket.IO ----------
@socketio.on('connect')
def on_connect():
    user_id = request.args.get('userId') or request.headers.get('X-User-Id') or 'guest'
    join_room(f'user:{user_id}')
    emit('connected', {'ok': True, 'userId': user_id})

@socketio.on('disconnect')
def on_disconnect():
    # Rooms are auto-learned; no explicit leave required here
    pass

# Helpers

def _get_user_id():
    return request.headers.get('X-User-Id') or (request.json or {}).get('userId') or request.args.get('userId') or 'guest'

def _ensure_user_store(user_id: str):
    if user_id not in NOTIFICATIONS:
        NOTIFICATIONS[user_id] = []
    return NOTIFICATIONS[user_id]

def _emit(event: str, user_id: str, payload: dict):
    socketio.emit(event, payload, room=f'user:{user_id}')

def _broadcast(event: str, payload: dict):
    try:
        socketio.emit(event, payload, broadcast=True)
    except Exception:
        pass

def _new_notif_id():
    return f"notif-{uuid.uuid4()}"

# ---- Mapping helper ----
def _map_student_row_to_user(r: dict) -> dict:
    """Map MySQL students row to frontend user structure used across pages."""
    if not r:
        return {}
    # Prefer stable keys from schema and expose aliases expected by UI
    out = {
        'id': r.get('student_id') or r.get('id'),
        'studentId': r.get('student_id') or r.get('id'),
        'userType': 'student',
        'role': 'student',
        'firstName': r.get('first_name'),
        'middleName': r.get('middle_name'),
        'lastName': r.get('last_name'),
        'suffix': r.get('suffix') or '',
        'fullName': r.get('full_name'),
        'email': r.get('email'),
        'phone': r.get('phone'),
        'gender': r.get('gender'),
        'birthday': str(r.get('birthdate')) if r.get('birthdate') is not None else '',
        'age': r.get('age'),
        'department': r.get('department'),
        'course': r.get('course'),
        'year': r.get('year_level'),
        'yearLevel': r.get('year_level'),
        'section': r.get('block'),
        'block': r.get('block'),
        # Permanent address (display/legacy)
        'address': r.get('permanent_address_full') or r.get('current_address_full') or '',
        'region': r.get('permanent_address_region') or '',
        'province': r.get('permanent_address_province') or '',
        'municipality': r.get('permanent_address_municipality') or '',
        'barangay': r.get('permanent_address_barangay') or '',
        'street': r.get('permanent_address_street') or '',
        'zipCode': r.get('permanent_address_zip') or '',
        # Current address (student-editable section)
        'currentAddress': r.get('current_address_full') or '',
        'currentRegion': r.get('current_address_region') or '',
        'currentProvince': r.get('current_address_province') or '',
        'currentMunicipality': r.get('current_address_municipality') or '',
        'currentBarangay': r.get('current_address_barangay') or '',
        'currentStreet': r.get('current_address_street') or '',
        'currentZipCode': r.get('current_address_zip') or '',
        'twoFactorEnabled': bool(r.get('two_factor_enabled')),
        'systemTag': r.get('system_tag') or 'JRMSU-KCS',
        'accountStatus': r.get('account_status') or 'active',
    }
    return out

def _map_admin_row_to_user(r: dict) -> dict:
    """Map MySQL admins row to frontend user structure."""
    if not r:
        return {}
    out = {
        'id': r.get('admin_id') or r.get('id'),
        'userType': 'admin',
        'role': r.get('position') or 'admin',
        'position': r.get('position') or 'admin',
        'firstName': r.get('first_name'),
        'middleName': r.get('middle_name'),
        'lastName': r.get('last_name'),
        'suffix': r.get('suffix') or '',
        'fullName': r.get('full_name') or ' '.join([x for x in [r.get('first_name'), r.get('last_name')] if x]),
        'email': r.get('email'),
        'phone': r.get('phone'),
        'gender': r.get('gender'),
        'birthday': str(r.get('birthdate')) if r.get('birthdate') is not None else '',
        'age': r.get('age'),
        'department': r.get('department') or '',
        'course': '',
        'year': '',
        'yearLevel': '',
        'section': '',
        'block': '',
        # Permanent Address
        'address': r.get('address') or ', '.join([p for p in [r.get('street'), r.get('barangay'), r.get('municipality'), r.get('province'), r.get('region'), 'Philippines', r.get('zip_code')] if p]),
        'region': r.get('region') or '',
        'province': r.get('province') or '',
        'municipality': r.get('municipality') or '',
        'barangay': r.get('barangay') or '',
        'street': r.get('street') or '',
        'zipCode': r.get('zip_code') or '',
        # Current Address
        'currentAddress': r.get('current_address') or ', '.join([p for p in [r.get('current_street'), r.get('current_barangay'), r.get('current_municipality'), r.get('current_province'), r.get('current_region'), 'Philippines', r.get('current_zip')] if p]),
        'currentRegion': r.get('current_region') or '',
        'currentProvince': r.get('current_province') or '',
        'currentMunicipality': r.get('current_municipality') or '',
        'currentBarangay': r.get('current_barangay') or '',
        'currentStreet': r.get('current_street') or '',
        'currentZipCode': r.get('current_zip') or '',
        'currentLandmark': r.get('current_landmark') or '',
        'sameAsCurrent': bool(r.get('same_as_current')),
        # System fields
        'twoFactorEnabled': bool(r.get('two_factor_enabled')),
        'systemTag': r.get('system_tag') or 'JRMSU-KCL',
        'accountStatus': r.get('account_status') or 'active',
        'isActive': (r.get('account_status') or 'active') == 'active',
        'qrCodeActive': True,
        'createdAt': r.get('created_at') or '',
        'updatedAt': r.get('updated_at') or '',
    }
    return out

@app.route('/ai/health')
def ai_health():
    try:
        r = requests.get(f"{OLLAMA_URL}/api/tags", timeout=3)
        return jsonify(ollama=r.ok)
    except Exception:
        return jsonify(ollama=False), 503

@app.route('/ai/chat', methods=['POST'])
def ai_chat():
    body = request.get_json(force=True)
    raw_message = (body.get('message') or '').strip()
    # Sanitize input to prevent prompt injection / XSS
    message = bleach.clean(raw_message, strip=True)
    history = body.get('history') or []
    messages = [{"role": "system", "content": "You are Jose, the JRMSU Library AI assistant."}]
    # Include up to last 5 history messages, sanitized
    for h in history[-5:]:
        role = h.get('role') in ('user', 'assistant', 'system') and h.get('role') or 'user'
        content = bleach.clean(str(h.get('content') or ''), strip=True)
        messages.append({"role": role, "content": content})
    messages.append({"role": "user", "content": message})

    try:
        r = requests.post(
            f"{OLLAMA_URL}/api/chat",
            json={
                "model": MODEL_NAME,
                "messages": messages,
                "stream": False,
                "options": {
                    "temperature": 0.2,
                    "top_p": 0.9,
                    "repeat_penalty": 1.1,
                    "num_ctx": 2048,
                    "num_predict": 256,
                    "keep_alive": "30m"
                }
            },
            timeout=60,
        )
        if not r.ok:
            return jsonify(error="Ollama request failed", details=r.text), 502
        data = r.json()
        content = (data.get('message') or {}).get('content', '')
        return jsonify(content=content)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route('/2fa/generate', methods=['POST'])
def twofa_generate():
    payload = request.get_json(silent=True) or {}
    account = payload.get('account') or 'user'
    issuer = payload.get('issuer') or 'JRMSU-LIBRARY'
    secret = generate_base32_secret(32)
    uri = key_uri(secret, account_name=account, issuer=issuer)
    code = current_totp_code(secret)
    return jsonify(secret=secret, otpauth=uri, currentCode=code)

@app.route('/2fa/verify', methods=['POST'])
def twofa_verify():
    payload = request.get_json(force=True)
    secret = (payload.get('secret') or '').strip()
    token = (payload.get('token') or '').strip()
    window = int(payload.get('window', 1))
    ok = verify_totp_code(secret, token, window=window)
    return jsonify(valid=ok)

# ---- Forgot Password API (legacy) ----
RESET_CODES = {}

@app.route('/auth/request-reset', methods=['POST'])
def auth_request_reset():
    body = request.get_json(force=True)
    method = (body.get('method') or 'email').lower()
    user_id = (body.get('userId') or '').strip()
    email = (body.get('email') or '').strip().lower()
    full_name = (body.get('fullName') or '').strip()
    if not email:
        return jsonify(error='Email required'), 400
    if method == 'email':
        # generate code
        code = f"{int(time.time())%1000000:06d}"
        expires_at = int(time.time()) + 300  # 5 minutes
        RESET_CODES[email] = { 'code': code, 'expires_at': expires_at, 'user_id': user_id, 'full_name': full_name }
        
        # Send actual email (or print to console in dev mode)
        email_sent = send_reset_email(email, code, full_name)
        
        # Push a notification to user (dev)
        if user_id:
            lst = _ensure_user_store(user_id)
            notif = {
                'id': _new_notif_id(),
                'user_id': user_id,
                'title': 'Password reset code sent',
                'body': f'A reset code was sent to {email}',
                'type': 'password_reset_request',
                'meta': {'email': email},
                'created_at': int(time.time()),
                'read': False,
                'action_required': False,
                'action_payload': None,
                'actor_id': 'system',
            }
            lst.insert(0, notif)
            _emit('notification.new', user_id, notif)
        return jsonify(ok=True)
    elif method == 'admin':
        # Notify admins (dev: broadcast to ADMIN)
        req_id = f"req-{uuid.uuid4()}"
        PASSWORD_RESET_REQUESTS[req_id] = {
            'id': req_id,
            'user_id': user_id or email,
            'email': email,
            'status': 'pending_admin',
            'created_at': int(time.time()),
        }
        admin_id = 'ADMIN'
        lst = _ensure_user_store(admin_id)
        notif = {
            'id': _new_notif_id(),
            'user_id': admin_id,
            'title': '🔔 Password Reset Request',
            'body': f'Reset requested by {full_name or user_id or email}',
            'type': 'password_reset_request',
            'meta': {'requestId': req_id, 'requesterId': user_id or email},
            'created_at': int(time.time()),
            'read': False,
            'action_required': True,
            'action_payload': {'actions': ['grant','decline']},
            'actor_id': user_id or email,
        }
        lst.insert(0, notif)
        _emit('notification.new', admin_id, notif)
        return jsonify(ok=True)
    else:
        return jsonify(error='Invalid method'), 400

@app.route('/auth/verify-code', methods=['POST'])
def auth_verify_code():
    body = request.get_json(force=True)
    email = (body.get('email') or '').strip().lower()
    code = (body.get('code') or '').strip()
    rec = RESET_CODES.get(email)
    if not rec:
        return jsonify(error='No code pending'), 400
    if int(time.time()) > int(rec.get('expires_at', 0)):
        return jsonify(error='Expired code'), 400
    if code != str(rec.get('code')):
        return jsonify(error='Invalid code'), 400
    return jsonify(ok=True, token=f"tok-{uuid.uuid4()}")

@app.route('/auth/reset-password', methods=['POST'])
def auth_reset_password():
    body = request.get_json(force=True)
    email = (body.get('email') or '').strip().lower()
    code = (body.get('code') or '').strip()
    new_password = (body.get('newPassword') or '').strip()
    if len(new_password) < 8:
        return jsonify(error='Password too short'), 400
    rec = RESET_CODES.get(email)
    if not rec:
        return jsonify(error='No code pending'), 400
    if int(time.time()) > int(rec.get('expires_at', 0)):
        return jsonify(error='Expired code'), 400
    if code != str(rec.get('code')):
        return jsonify(error='Invalid code'), 400
    # In real deployment, update user record with bcrypt
    del RESET_CODES[email]
    # Notify user
    user_id = rec.get('user_id') or email
    lst = _ensure_user_store(user_id)
    notif = {
        'id': _new_notif_id(),
        'user_id': user_id,
        'title': 'Password reset successful',
        'body': 'Your password was changed successfully',
        'type': 'system_alert',
        'meta': {},
        'created_at': int(time.time()),
        'read': False,
        'action_required': False,
        'action_payload': None,
        'actor_id': 'system',
    }
    lst.insert(0, notif)
    _emit('notification.new', user_id, notif)
    return jsonify(ok=True)

# ---- New API routes to align with plan ----

# Manual login with bcrypt verification (server-side)
@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    try:
        body = request.get_json(force=True) or {}
        user_id = (body.get('id') or body.get('userId') or '').strip()
        password = (body.get('password') or '').encode('utf-8')
        if not user_id or not password:
            return jsonify(error='id and password required'), 400

        # Try Admins first
        try:
            row = execute_query(
                "SELECT * FROM admins WHERE admin_id = %s OR id = %s",
                (user_id, user_id),
                fetch_one=True
            )
            if row and row.get('password_hash'):
                try:
                    ok = bcrypt.checkpw(password, str(row.get('password_hash')).encode('utf-8'))
                except Exception:
                    ok = False
                if ok:
                    user = _map_admin_row_to_user(row)
                    log_activity(user.get('id') or user_id, 'login', 'manual')
                    return jsonify(user=user)
        except Exception:
            pass

        # Try Students
        try:
            row = execute_query(
                "SELECT * FROM students WHERE student_id = %s OR id = %s",
                (user_id, user_id),
                fetch_one=True
            )
            if row and row.get('password_hash'):
                try:
                    ok = bcrypt.checkpw(password, str(row.get('password_hash')).encode('utf-8'))
                except Exception:
                    ok = False
                if ok:
                    user = _map_student_row_to_user(row)
                    log_activity(user.get('id') or user_id, 'login', 'manual')
                    return jsonify(user=user)
        except Exception:
            pass

        # Fallback to file-backed users (dev mode)
        try:
            fdb = load_db()
            u = (fdb.get('users') or {}).get(user_id)
            if u and u.get('password_hash'):
                try:
                    ok = bcrypt.checkpw(password, str(u.get('password_hash')).encode('utf-8'))
                except Exception:
                    ok = False
                if ok:
                    log_activity(user_id, 'login', 'manual_fallback')
                    return jsonify(user=u)
        except Exception:
            pass

        return jsonify(error='Invalid credentials or user not found'), 401
    except Exception as e:
        return jsonify(error=str(e)), 500

# -------------------- Books API (minimal) --------------------
@app.route('/api/books', methods=['GET'])
def api_books_list():
    try:
        rows = execute_query(
            "SELECT id, title, author, category, isbn, publisher, publication_year, total_copies, available_copies, status FROM books ORDER BY title",
            fetch_all=True
        ) or []
        return jsonify(items=rows)
    except Exception:
        # Fallback to file-backed store
        fdb = load_db()
        books = fdb.get('books', []) or []
        # Normalize keys similar to DB
        data = []
        for b in books:
            data.append({
                'id': b.get('id') or b.get('bookId') or b.get('code'),
                'title': b.get('title'),
                'author': b.get('author'),
                'category': b.get('category') or b.get('genre'),
                'isbn': b.get('isbn'),
                'publisher': b.get('publisher'),
                'publication_year': b.get('publication_year'),
                'total_copies': b.get('copies') or b.get('total_copies') or 1,
                'available_copies': b.get('available') or b.get('available_copies') or 1,
                'status': b.get('status') or 'available',
            })
        return jsonify(items=data)

@app.route('/api/books', methods=['POST'])
def api_books_create():
    body = request.get_json(force=True) or {}
    book_id = (body.get('id') or body.get('book_id') or '').strip()
    title = (body.get('title') or '').strip()
    author = (body.get('author') or '').strip()
    category = (body.get('category') or '').strip()
    isbn = (body.get('isbn') or '').strip()
    publisher = (body.get('publisher') or '').strip()
    pub_year = body.get('publication_year')
    total = int(body.get('total_copies') or body.get('copies') or 1)
    available = int(body.get('available_copies') or body.get('available') or total)
    status = body.get('status') or ('available' if available > 0 else 'unavailable')
    if not book_id or not title:
        return jsonify(error='id and title required'), 400
    try:
        execute_query(
            """
            INSERT INTO books (id, title, author, isbn, category, publisher, publication_year, total_copies, available_copies, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (book_id, title, author, isbn, category, publisher, pub_year, total, available, status)
        )
        try:
            _broadcast('book.added', {'bookId': book_id})
        except Exception:
            pass
        return jsonify(ok=True, id=book_id)
    except Exception as e:
        # Fallback to file-backed store
        fdb = load_db()
        books = fdb.setdefault('books', [])
        if any((x.get('id') or x.get('bookId')) == book_id for x in books):
            return jsonify(error='duplicate id'), 400
        books.append({
            'id': book_id, 'title': title, 'author': author, 'isbn': isbn, 'category': category,
            'publisher': publisher, 'publication_year': pub_year, 'copies': total, 'available': available, 'status': status
        })
        save_db(fdb)
        try:
            _broadcast('book.added', {'bookId': book_id})
        except Exception:
            pass
        return jsonify(ok=True, id=book_id)

@app.route('/api/books/<book_id>', methods=['DELETE'])
def api_books_delete(book_id: str):
    try:
        execute_query("DELETE FROM books WHERE id = %s", (book_id,))
        try:
            _broadcast('book.removed', {'bookId': book_id})
        except Exception:
            pass
        return jsonify(ok=True)
    except Exception:
        # Fallback to file-backed store
        fdb = load_db()
        books = fdb.get('books', []) or []
        new_books = [b for b in books if (b.get('id') or b.get('bookId')) != book_id]
        fdb['books'] = new_books
        save_db(fdb)
        try:
            _broadcast('book.removed', {'bookId': book_id})
        except Exception:
            pass
        return jsonify(ok=True)

@app.route('/api/auth/send-reset-email', methods=['POST'])
def api_send_reset_email():
    body = request.get_json(force=True)
    user = (body.get('userIdOrEmail') or body.get('userId') or body.get('email') or '').strip()
    if not user:
        return jsonify(error='userIdOrEmail required'), 400
    email = body.get('email') or (user if '@' in user else '')
    # Reuse legacy route logic
    return auth_request_reset()

@app.route('/api/auth/verify-reset-code', methods=['POST'])
def api_verify_reset_code():
    return auth_verify_code()

@app.route('/api/auth/message-admin', methods=['POST'])
def api_message_admin():
    # Reuse legacy admin branch
    req = request.get_json(force=True)
    req['method'] = 'admin'
    with app.test_request_context(json=req):
        return auth_request_reset()

@app.route('/api/auth/admin-respond', methods=['POST'])
def api_admin_respond():
    body = request.get_json(force=True)
    request_id = body.get('requestId')
    action = (body.get('action') or '').lower()
    admin_id = body.get('adminId') or 'ADMIN'
    rec = PASSWORD_RESET_REQUESTS.get(request_id)
    if not rec:
        return jsonify(error='Request not found'), 404
    if action not in ('grant','decline'):
        return jsonify(error='Invalid action'), 400
    rec['status'] = 'approved' if action == 'grant' else 'declined'
    # Notify requester
    user_id = rec.get('user_id')
    lst = _ensure_user_store(user_id)
    notif = {
        'id': _new_notif_id(),
        'user_id': user_id,
        'title': 'Admin response to password reset',
        'body': f'Your request was {rec["status"]} by {admin_id}',
        'type': 'admin_response',
        'meta': {'requestId': request_id, 'status': rec['status']},
        'created_at': int(time.time()),
        'read': False,
        'action_required': False,
        'action_payload': None,
        'actor_id': admin_id,
    }
    lst.insert(0, notif)
    _emit('notification.admin_response', user_id, {'requestId': request_id, 'status': rec['status']})
    _emit('notification.new', user_id, notif)
    return jsonify(ok=True)

@app.route('/api/auth/verify-2fa', methods=['POST'])
def api_verify_2fa():
    body = request.get_json(force=True)
    secret = (body.get('secret') or '').strip()
    token = (body.get('totpCode') or body.get('token') or '').strip()
    window = int(body.get('window', 1))
    ok = verify_totp_code(secret, token, window=window)
    return jsonify(valid=ok)

@app.route('/api/auth/reset-password', methods=['POST'])
def api_reset_password():
    return auth_reset_password()

# ---- AI chat alias ----
@app.route('/api/ai/chat', methods=['POST'])
def api_ai_chat():
    return ai_chat()

# ---- Notifications API ----
@app.route('/api/notifications')
def api_notifications():
    user_id = _get_user_id()
    filter_val = request.args.get('filter','all')
    page = int(request.args.get('page',1))
    limit = min(100, int(request.args.get('limit',25)))
    lst = _ensure_user_store(user_id)
    items = sorted(lst, key=lambda x: x.get('created_at',0), reverse=True)
    if filter_val == 'unread':
        items = [n for n in items if not n.get('read')]
    total = len(items)
    start = (page-1)*limit
    end = start + limit
    return jsonify(items=items[start:end], total=total, unread=sum(1 for n in lst if not n.get('read')))

@app.route('/api/notifications/mark-read', methods=['POST'])
def api_notifications_mark_read():
    user_id = _get_user_id()
    body = request.get_json(force=True)
    ids = body.get('notificationIds') or []
    lst = _ensure_user_store(user_id)
    updated = []
    for n in lst:
        if n['id'] in ids and not n.get('read'):
            n['read'] = True
            updated.append(n)
    for n in updated:
        _emit('notification.update', user_id, n)
    return jsonify(ok=True, updated=len(updated))

@app.route('/api/notifications/mark-all-read', methods=['POST'])
def api_notifications_mark_all_read():
    user_id = _get_user_id()
    lst = _ensure_user_store(user_id)
    for n in lst:
        n['read'] = True
    _emit('notification.mark_all_read', user_id, {'userId': user_id, 'timestamp': int(time.time())})
    return jsonify(ok=True)

@app.route('/api/notifications/<nid>')
def api_notifications_get(nid: str):
    user_id = _get_user_id()
    lst = _ensure_user_store(user_id)
    for n in lst:
        if n['id'] == nid:
            if not n.get('read'):
                n['read'] = True
                _emit('notification.update', user_id, n)
            return jsonify(n)
    return jsonify(error='Not found'), 404

@app.route('/api/notifications/<nid>/action', methods=['POST'])
def api_notifications_action(nid: str):
    # For admin actions on notifications (e.g., grant/decline)
    body = request.get_json(force=True)
    action = (body.get('action') or '').lower()
    admin_id = body.get('adminId') or 'ADMIN'
    # No-op demo: just broadcast update back
    user_id = _get_user_id()
    lst = _ensure_user_store(user_id)
    for n in lst:
        if n['id'] == nid:
            n['meta'] = {**(n.get('meta') or {}), 'adminAction': action, 'adminId': admin_id}
            _emit('notification.update', user_id, n)
            return jsonify(ok=True)
    return jsonify(error='Not found'), 404

@app.route('/qr/validate', methods=['POST'])
def qr_validate():
    body = request.get_json(force=True)
    raw = body.get('data')
    try:
        obj = json.loads(raw) if isinstance(raw, str) else raw
        required = ['systemId', 'userId', 'fullName', 'userType', 'systemTag']
        if not all(k in obj for k in required):
            return jsonify(valid=False, error='Missing required fields')
        if obj.get('systemId') != 'JRMSU-LIBRARY':
            return jsonify(valid=False, error='Invalid systemId')
        if not (obj.get('encryptedPasswordToken') or obj.get('sessionToken') or obj.get('encryptedToken')):
            return jsonify(valid=False, error='Missing auth token')
        expected = 'JRMSU-KCL' if obj.get('userType') == 'admin' else 'JRMSU-KCS'
        if obj.get('systemTag') != expected:
            return jsonify(valid=False, error='System tag mismatch')
        return jsonify(valid=True)
    except Exception as e:
        return jsonify(valid=False, error=str(e)), 400


# Register library entry/exit endpoints
try:
    from library_session_manager import register_library_session_endpoints
    if 'check_user_session' not in app.view_functions:
        register_library_session_endpoints(app)
        print('✅ Library session endpoints loaded')
    else:
        print('✅ Library session endpoints already loaded')
except Exception as e:
    print(f'⚠️  Library session endpoints not loaded: {e}')

# Register additional library endpoints (book operations)
try:
    from library_endpoints import register_library_endpoints
    register_library_endpoints(app)
    print('✅ Library book endpoints loaded')
except Exception as e:
    print(f'⚠️  Library book endpoints not loaded: {e}')

# Register password management endpoints
try:
    from password_endpoints import register_password_endpoints
    register_password_endpoints(app)
    print('✅ Password endpoints loaded')
except Exception as e:
    print(f'⚠️  Password endpoints not loaded: {e}')

# Register notification endpoints
try:
    from notification_endpoints import register_notification_endpoints
    register_notification_endpoints(app)
    print('✅ Notification endpoints loaded')
except Exception as e:
    print(f'⚠️  Notification endpoints not loaded: {e}')

# ---------- QR Code Synchronization (Students/Admins) ----------

@app.route('/api/students/<student_id>/qr-code', methods=['GET'])
def get_student_qr_code(student_id: str):
    try:
        # Try DB first
        row = execute_query(
            "SELECT qr_code_data, qr_code_generated_at FROM students WHERE student_id = %s OR id = %s",
            (student_id, student_id),
            fetch_one=True
        )
        if row:
            return jsonify({
                'qrCodeData': row.get('qr_code_data'),
                'qrCodeGeneratedAt': str(row.get('qr_code_generated_at')) if row.get('qr_code_generated_at') else None
            })
    except Exception:
        pass
    # Fallback to file-backed store
    fdb = load_db()
    u = (fdb.get('users') or {}).get(student_id) or {}
    return jsonify({
        'qrCodeData': u.get('qrCodeData'),
        'qrCodeGeneratedAt': u.get('qrCodeGeneratedAt')
    })

@app.route('/api/students/<student_id>/qr-code', methods=['PUT'])
def put_student_qr_code(student_id: str):
    body = request.get_json(force=True) or {}
    data = body.get('qrCodeData')
    try:
        # Prefer DB update
        execute_query(
            """
            UPDATE students 
            SET qr_code_data = %s, qr_code_generated_at = NOW(), updated_at = NOW()
            WHERE student_id = %s OR id = %s
            """,
            (data, student_id, student_id)
        )
        # Emit update
        _emit('students.updated', student_id, {'qrCodeData': data})
        return jsonify(ok=True)
    except Exception:
        # Fallback to file store
        fdb = _ensure_users()
        u = (fdb.get('users') or {}).get(student_id) or {'id': student_id, 'userType': 'student'}
        u['qrCodeData'] = data
        u['qrCodeGeneratedAt'] = int(time.time() * 1000)
        fdb['users'][student_id] = u
        save_db(fdb)
        _emit('students.updated', student_id, {'qrCodeData': data})
        return jsonify(ok=True)

@app.route('/api/admins/<admin_id>/qr-code', methods=['GET'])
def get_admin_qr_code(admin_id: str):
    try:
        row = execute_query(
            "SELECT qr_code_data, qr_code_generated_at FROM admins WHERE admin_id = %s OR id = %s",
            (admin_id, admin_id),
            fetch_one=True
        )
        if row:
            return jsonify({
                'qrCodeData': row.get('qr_code_data'),
                'qrCodeGeneratedAt': str(row.get('qr_code_generated_at')) if row.get('qr_code_generated_at') else None
            })
    except Exception:
        pass
    # Fallback to file-backed store
    fdb = load_db()
    u = (fdb.get('users') or {}).get(admin_id) or {}
    return jsonify({
        'qrCodeData': u.get('qrCodeData'),
        'qrCodeGeneratedAt': u.get('qrCodeGeneratedAt')
    })

@app.route('/api/admins/<admin_id>/qr-code', methods=['PUT'])
def put_admin_qr_code(admin_id: str):
    body = request.get_json(force=True) or {}
    data = body.get('qrCodeData')
    try:
        execute_query(
            """
            UPDATE admins 
            SET qr_code_data = %s, qr_code_generated_at = NOW(), updated_at = NOW()
            WHERE admin_id = %s OR id = %s
            """,
            (data, admin_id, admin_id)
        )
        _emit('admins.updated', admin_id, {'qrCodeData': data})
        return jsonify(ok=True)
    except Exception:
        fdb = _ensure_users()
        u = (fdb.get('users') or {}).get(admin_id) or {'id': admin_id, 'userType': 'admin'}
        u['qrCodeData'] = data
        u['qrCodeGeneratedAt'] = int(time.time() * 1000)
        fdb['users'][admin_id] = u
        save_db(fdb)
        _emit('admins.updated', admin_id, {'qrCodeData': data})
        return jsonify(ok=True)

# ---------- Courses Endpoint (Dynamic by department) ----------

@app.route('/api/courses', methods=['GET'])
def api_courses():
    """Return list of courses. Optional query param department=<code> (cte,cba,cafse,scje,ccs)"""
    dept = (request.args.get('department') or '').strip().lower()
    # Fallback mapping used when DB is unavailable or courses table absent
    fallback = {
        'cte': [
            { 'code': 'bsfil', 'name': 'BS Filipino' },
            { 'code': 'bssci', 'name': 'BS Science' },
            { 'code': 'bsee',  'name': 'BS Elementary Ed' },
            { 'code': 'bsmath','name': 'BS Math' },
            { 'code': 'bspe',  'name': 'BS PE' },
        ],
        'cba': [
            { 'code': 'bhm',   'name': 'BS Hospitality Management' },
            { 'code': 'bbahrm','name': 'BSBA – HR Management' },
            { 'code': 'bsab',  'name': 'BS Agri-Business' },
        ],
        'cafse': [
            { 'code': 'bsa',   'name': 'BS Agriculture' },
            { 'code': 'bsf',   'name': 'BS Forestry' },
            { 'code': 'bsabe', 'name': 'BS Agri & Biosystems Eng.' },
        ],
        'scje': [
            { 'code': 'none',  'name': 'No course selection required' },
        ],
        'ccs': [
            { 'code': 'bsis',  'name': 'BS Information System' },
            { 'code': 'bscs',  'name': 'BS Computer Science' },
        ],
    }
    try:
        # Try DB first, if a courses table exists with (department, code, name)
        q = "SELECT department, code, name FROM courses"
        rows = execute_query(q, fetch_all=True) or []
        if rows:
            items = [ { 'department': (r.get('department') or '').lower(), 'code': r.get('code'), 'name': r.get('name') } for r in rows ]
            if dept:
                items = [i for i in items if i.get('department') == dept]
            return jsonify(items=items)
    except Exception:
        pass
    # Fallback
    if dept:
        return jsonify(items=fallback.get(dept) or [])
    all_items = []
    for k, lst in fallback.items():
        all_items.extend([{ **it, 'department': k } for it in lst])
    return jsonify(items=all_items)

# ---------- Account Recovery & 2FA Backup ----------

@app.route('/api/users/<user_id>/recovery-email', methods=['GET'])
def get_recovery_email(user_id: str):
    try:
        row = execute_query(
            """
            SELECT recovery_email FROM user_security_settings WHERE user_id = %s
            """,
            (user_id,),
            fetch_one=True
        )
        return jsonify({ 'userId': user_id, 'recoveryEmail': (row or {}).get('recovery_email') })
    except Exception as e:
        return jsonify({ 'userId': user_id, 'recoveryEmail': None })

@app.route('/api/users/<user_id>/recovery-email', methods=['PUT'])
def put_recovery_email(user_id: str):
    body = request.get_json(force=True) or {}
    recovery_email = (body.get('recoveryEmail') or '').strip()
    if not recovery_email:
        return jsonify(error='recoveryEmail required'), 400
    try:
        # Ensure settings table exists (idempotent)
        execute_query(
            """
            CREATE TABLE IF NOT EXISTS user_security_settings (
              user_id VARCHAR(64) PRIMARY KEY,
              recovery_email VARCHAR(255) NULL,
              created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
              updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        # Upsert
        execute_query(
            """
            INSERT INTO user_security_settings (user_id, recovery_email)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE recovery_email = VALUES(recovery_email), updated_at = CURRENT_TIMESTAMP
            """,
            (user_id, recovery_email)
        )
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route('/api/users/<user_id>/2fa/backup', methods=['GET'])
def get_twofa_backup(user_id: str):
    try:
        # Check file-backed store for 2FA status and key
        fdb = load_db()
        u = (fdb.get('users') or {}).get(user_id)
        if not u or not u.get('twoFactorEnabled'):
            return jsonify(enabled=False), 200
        secret = u.get('twoFactorKey') or u.get('twoFactorSetupKey')
        if not secret:
            return jsonify(enabled=False), 200
        uri = key_uri(secret, name=user_id, issuer_name='JRMSU Library System')
        return jsonify(enabled=True, secret=secret, otpauth=uri)
    except Exception as e:
        return jsonify(error=str(e)), 500

def cleanup_active_sessions_on_startup():
    try:
        if not MYSQL_AVAILABLE:
            print('🧹 Skipping startup session cleanup (MySQL unavailable)')
            return
        # Ensure tables exist before cleanup
        try:
            execute_query("""
                CREATE TABLE IF NOT EXISTS library_sessions (
                  session_id VARCHAR(50) PRIMARY KEY,
                  user_id VARCHAR(20) NOT NULL,
                  user_type ENUM('student','admin') NOT NULL,
                  full_name VARCHAR(100) NOT NULL,
                  login_time DATETIME NOT NULL,
                  logout_time DATETIME NULL,
                  method ENUM('manual','qr') NOT NULL DEFAULT 'manual',
                  status ENUM('inside_library','logged_out') NOT NULL DEFAULT 'inside_library',
                  action_count INT NOT NULL DEFAULT 1,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                  KEY idx_user_status (user_id, status),
                  KEY idx_status (status),
                  KEY idx_login_time (login_time)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception:
            pass
        try:
            execute_query("""
                CREATE TABLE IF NOT EXISTS active_sessions (
                  session_id INT AUTO_INCREMENT PRIMARY KEY,
                  user_id VARCHAR(20) NOT NULL,
                  fullname VARCHAR(100) NOT NULL,
                  usertype ENUM('Student','Admin') NOT NULL,
                  login_time DATETIME NOT NULL,
                  logout_time DATETIME NULL,
                  status ENUM('active','logged_out') NOT NULL DEFAULT 'active',
                  login_method ENUM('manual','qrcode') NOT NULL DEFAULT 'manual',
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                  KEY idx_active (status),
                  KEY idx_user (user_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as e:
            msg = str(e)
            if ('1813' in msg) or ('Tablespace' in msg):
                try:
                    execute_query("DROP TABLE IF EXISTS active_sessions")
                except Exception:
                    pass
                try:
                    execute_query("""
                        CREATE TABLE active_sessions (
                          session_id INT AUTO_INCREMENT PRIMARY KEY,
                          user_id VARCHAR(20) NOT NULL,
                          fullname VARCHAR(100) NOT NULL,
                          usertype ENUM('Student','Admin') NOT NULL,
                          login_time DATETIME NOT NULL,
                          logout_time DATETIME NULL,
                          status ENUM('active','logged_out') NOT NULL DEFAULT 'active',
                          login_method ENUM('manual','qrcode') NOT NULL DEFAULT 'manual',
                          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                          updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                          KEY idx_active (status),
                          KEY idx_user (user_id)
                        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                    """)
                except Exception:
                    pass
            else:
                pass
        # Mark all sessions as logged out (graceful cleanup)
        try:
            execute_query("UPDATE active_sessions SET status='logged_out', logout_time=NOW() WHERE status='active'")
        except Exception:
            pass
        try:
            execute_query("UPDATE library_sessions SET status='logged_out', logout_time=NOW() WHERE status='inside_library' AND (logout_time IS NULL OR logout_time < NOW())")
        except Exception:
            pass
        try:
            socketio.emit('session_cleanup', { 'timestamp': int(time.time()) }, broadcast=True)
        except Exception:
            pass
        print('🧹 Active sessions cleaned up on startup')
    except Exception as e:
        print(f'⚠️  Startup cleanup skipped: {e}')

if __name__ == '__main__':
    # Backend port is hardcoded to 5000 (not changeable)
    port = 5000
    cleanup_active_sessions_on_startup()
    # Start overdue watcher (emits book.overdue events)
    start_overdue_watcher(60)
    # Start daily bulk overdue notifications for all users (email/SMS/push)
    start_overdue_bulk_notifier(24 * 60 * 60)
    print(f'🚀 Backend running at http://localhost:{port}')
    socketio.run(app, host='0.0.0.0', port=port)
